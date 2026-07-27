import os
import json
import pandas as pd
import re
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if not os.path.exists(BASE_DIR) or os.path.basename(BASE_DIR) != 'dashboard.pbjkobar-testing':
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))

tahun_n  = datetime.now().year

SET_PEMILIHAN = {'bapbast', 'berlangsung', 'completed', 'kontrak', 'non tender selesai', 'on process', 'paket sedang berjalan', 'paket selesai', 'payment outside system', 'spmkspp', 'sppbj'}
SET_KONTRAK = {'bapbast', 'completed', 'kontrak', 'on process', 'paket sedang berjalan', 'paket selesai', 'payment outside system', 'spmkspp'}
SET_SERAH_TERIMA = {'bapbast', 'completed', 'paket selesai', 'payment outside system'}

def kelola_arsip_bulanan(folder_path, tahun):
    if not os.path.exists(folder_path): return
    # Arahkan ke folder arsip_lokal/tepra
    folder_arsip_lokal = os.path.join(BASE_DIR, 'arsip_lokal', 'tepra', str(tahun))
    os.makedirs(folder_arsip_lokal, exist_ok=True)
    
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    arsip_bulanan = {}
    for f in file_excel:
        match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", f)
        if match:
            bulan = match.group(1)[:7]
            if bulan not in arsip_bulanan: arsip_bulanan[bulan] = []
            arsip_bulanan[bulan].append((match.group(1), f))
            
    for bulan, list_file in arsip_bulanan.items():
        list_file.sort(key=lambda x: x[0])
        # Pindahkan semua file kecuali yang paling baru (terakhir) di bulan tersebut
        for tgl, nama_file in list_file[:-1]:
            try: shutil.move(os.path.join(folder_path, nama_file), os.path.join(folder_arsip_lokal, nama_file))
            except: pass

def update_daftar_arsip_json(folder_path):
    if not os.path.exists(folder_path): return
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    file_excel.sort(reverse=True) 
    arsip_list = [{"nama_file": f} for f in file_excel]
    try:
        with open(os.path.join(folder_path, 'daftar_arsip.json'), 'w', encoding='utf-8') as f:
            json.dump(arsip_list, f, indent=4)
    except: pass

def safe_float(val):
    try: return float(str(val).replace(',', ''))
    except: return 0.0

def get_data_rekap(df_source, min_pagu, max_pagu):
    # Ambil daftar seluruh Satuan Kerja dari sumber utama agar tidak ada yang terlewat (Pasti 58)
    semua_satker = sorted(df_source['Satuan Kerja'].astype(str).str.strip().unique())
    
    # Filter data sesuai rentang Pagu
    df = df_source[(df_source['Nilai Pagu RUP'] > min_pagu) & (df_source['Nilai Pagu RUP'] <= max_pagu)].copy()
    
    if not df.empty:
        df['Status_Clean'] = df['Status'].astype(str).str.strip().str.lower()
        df['Satuan Kerja'] = df['Satuan Kerja'].astype(str).str.strip()
        grouped = df.groupby('Satuan Kerja')
    else:
        grouped = None

    rekap = []
    # Looping berdasarkan SEMUA SATUAN KERJA mutlak, bukan berdasarkan data yang terfilter
    for satker in semua_satker:
        if grouped is not None and satker in grouped.groups:
            group = grouped.get_group(satker)
            mask_pemilihan = group['Status_Clean'].isin(SET_PEMILIHAN)
            paket_pemilihan = mask_pemilihan.sum()
            pagu_pemilihan = group.loc[mask_pemilihan, 'Nilai Pagu RUP'].sum()
            hasil_pemilihan = group.loc[mask_pemilihan, 'Nilai Hasil Pemilihan'].sum()

            mask_kontrak = group['Status_Clean'].isin(SET_KONTRAK)
            paket_kontrak = mask_kontrak.sum()
            pagu_kontrak = group.loc[mask_kontrak, 'Nilai Hasil Pemilihan'].sum()

            mask_serah_terima = group['Status_Clean'].isin(SET_SERAH_TERIMA)
            paket_serah_terima = mask_serah_terima.sum()
            pagu_serah_terima = group.loc[mask_serah_terima, 'Nilai Hasil Pemilihan'].sum()

            mask_belum = ~mask_pemilihan
            paket_belum = mask_belum.sum()
            pagu_belum = group.loc[mask_belum, 'Nilai Pagu RUP'].sum()
        else:
            # Jika Satker tidak punya paket di rentang uang ini, isi dengan 0
            paket_pemilihan = pagu_pemilihan = hasil_pemilihan = 0
            paket_kontrak = pagu_kontrak = 0
            paket_serah_terima = pagu_serah_terima = 0
            paket_belum = pagu_belum = 0

        rekap.append({
            'Satuan Kerja': satker,
            'Data Laporan_Paket': int(paket_pemilihan + paket_belum),
            'Data Laporan_Pagu': pagu_pemilihan + pagu_belum,
            'Pemilihan_Paket': int(paket_pemilihan),
            'Pemilihan_Pagu': pagu_pemilihan,
            'Hasil_Paket': int(paket_pemilihan),
            'Hasil_Pagu': hasil_pemilihan,
            'Kontrak_Paket': int(paket_kontrak),
            'Kontrak_Pagu': pagu_kontrak,
            'SerahTerima_Paket': int(paket_serah_terima),
            'SerahTerima_Pagu': pagu_serah_terima,
            'Belum_Paket': int(paket_belum),
            'Belum_Pagu': pagu_belum
        })

    df_rekap = pd.DataFrame(rekap).sort_values('Satuan Kerja')
    total_row = {'Satuan Kerja': 'TOTAL KESELURUHAN'}
    for col in df_rekap.columns:
        if col != 'Satuan Kerja': 
            total_row[col] = df_rekap[col].sum()
            
    return pd.concat([df_rekap, pd.DataFrame([total_row])], ignore_index=True)


# =======================================================================================
# FUNGSI STYLE EXCEL KHUSUS: 0 SD 50 JUTA (10 Kolom)
# =======================================================================================
def style_excel_sheet_0_50(ws, df_rekap, title_text):
    font_normal = Font(name='Arial', size=9)
    header_font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    header_font_italic = Font(name='Arial', bold=True, italic=True, size=9, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1F4E79')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Judul
    ws.merge_cells('A1:J1'); ws.merge_cells('A2:J2')
    ws['A1'] = title_text
    bln = {1:'JANUARI', 2:'FEBRUARI', 3:'MARET', 4:'APRIL', 5:'MEI', 6:'JUNI', 7:'JULI', 8:'AGUSTUS', 9:'SEPTEMBER', 10:'OKTOBER', 11:'NOVEMBER', 12:'DESEMBER'}
    ws['A2'] = f"KABUPATEN KOTAWARINGIN BARAT S/D {datetime.now().day} {bln[datetime.now().month]} {datetime.now().year}"
    ws['A1'].font = ws['A2'].font = Font(name='Arial', bold=True, size=11)
    ws['A1'].alignment = ws['A2'].alignment = align_center

    # Header Tabel
    ws.merge_cells('A4:A7'); ws['A4'] = "NO"
    ws.merge_cells('B4:B7'); ws['B4'] = "SATUAN KERJA"
    ws.merge_cells('C4:D5'); ws['C4'] = "DATA LAPORAN"
    ws.merge_cells('E4:F5'); ws['E4'] = "DATA REALISASI (RUP)"
    ws.merge_cells('G4:H5'); ws['G4'] = "DATA REALISASI"
    ws.merge_cells('I4:J5'); ws['I4'] = "BELUM PENGADAAN"
    
    ws.merge_cells('C6:C7'); ws['C6'] = "JUMLAH\nPAKET"
    ws.merge_cells('D6:D7'); ws['D6'] = "JUMLAH PAGU\n(Rp)"
    ws.merge_cells('E6:E7'); ws['E6'] = "JUMLAH\nPAKET"
    ws.merge_cells('F6:F7'); ws['F6'] = "JUMLAH PAGU\n(Rp)"
    ws.merge_cells('G6:G7'); ws['G6'] = "JUMLAH\nPAKET"
    ws.merge_cells('H6:H7'); ws['H6'] = "JUMLAH PAGU\n(Rp)"
    ws.merge_cells('I6:I7'); ws['I6'] = "JUMLAH\nPAKET"
    ws.merge_cells('J6:J7'); ws['J6'] = "JUMLAH PAGU\n(Rp)"
    ws.row_dimensions[6].height = 30

    numbers = ['1','2','3','4','5','6','7','8','9','10']
    for col_idx, val in enumerate(numbers, start=1): ws.cell(row=8, column=col_idx, value=val)

    for r in range(4, 9):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font_italic if r == 8 else header_font
            cell.fill, cell.alignment, cell.border = header_fill, align_center, border_thin

    if df_rekap.empty: return

    # Isi Data
    row_num = 9
    for idx, row in df_rekap.iterrows():
        is_total = row['Satuan Kerja'] == 'TOTAL KESELURUHAN'
        data_row = [
            "" if is_total else row_num - 8, row['Satuan Kerja'],
            row['Data Laporan_Paket'] if row['Data Laporan_Paket'] else "-", row['Data Laporan_Pagu'] if row['Data Laporan_Pagu'] else "-",
            row['Pemilihan_Paket'] if row['Pemilihan_Paket'] else "-", row['Pemilihan_Pagu'] if row['Pemilihan_Pagu'] else "-",
            row['Hasil_Paket'] if row['Hasil_Paket'] else "-", row['Hasil_Pagu'] if row['Hasil_Pagu'] else "-",
            row['Belum_Paket'] if row['Belum_Paket'] else "-", row['Belum_Pagu'] if row['Belum_Pagu'] else "-"
        ]
        
        for col_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if is_total:
                cell.font, cell.fill = header_font, header_fill
            else:
                cell.font = font_normal
                cell.fill = PatternFill('solid', start_color='FFFFFF') if row_num % 2 != 0 else PatternFill('solid', start_color='DCE6F1')
            
            cell.border = border_thin
            if col_idx == 2: cell.alignment = align_center if is_total else align_left
            elif col_idx in [4, 6, 8, 10] and isinstance(val, (int, float)):
                cell.number_format, cell.alignment = '#,##0', align_right
            else:
                cell.alignment = align_center
                if isinstance(val, (int, float)): cell.number_format = '#,##0'
                
        ws.row_dimensions[row_num].height = 45
        row_num += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col in ['C', 'E', 'G', 'I']: ws.column_dimensions[col].width = 8
    for col in ['D', 'F', 'H', 'J']: ws.column_dimensions[col].width = 16


# =======================================================================================
# FUNGSI STYLE EXCEL STANDAR: (14 Kolom)
# =======================================================================================
def style_excel_sheet(ws, df_rekap, title_text):
    font_normal = Font(name='Arial', size=9)
    header_font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    header_font_italic = Font(name='Arial', bold=True, italic=True, size=9, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1F4E79')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Judul
    ws.merge_cells('A1:N1'); ws.merge_cells('A2:N2')
    ws['A1'] = title_text
    bln = {1:'JANUARI', 2:'FEBRUARI', 3:'MARET', 4:'APRIL', 5:'MEI', 6:'JUNI', 7:'JULI', 8:'AGUSTUS', 9:'SEPTEMBER', 10:'OKTOBER', 11:'NOVEMBER', 12:'DESEMBER'}
    ws['A2'] = f"KABUPATEN KOTAWARINGIN BARAT S/D {datetime.now().day} {bln[datetime.now().month]} {datetime.now().year}"
    ws['A1'].font = ws['A2'].font = Font(name='Arial', bold=True, size=11)
    ws['A1'].alignment = ws['A2'].alignment = align_center

    # Header Tabel
    ws.merge_cells('A4:A7'); ws['A4'] = "NO"
    ws.merge_cells('B4:B7'); ws['B4'] = "SATUAN KERJA"
    ws.merge_cells('C4:D4'); ws['C4'] = "DATA LAPORAN"
    ws.merge_cells('E4:L4'); ws['E4'] = "PROSES PENGADAAN"
    ws.merge_cells('M4:N4'); ws['M4'] = "BELUM PENGADAAN"
    ws.merge_cells('C5:C7'); ws['C5'] = "JUMLAH\nPAKET"
    ws.merge_cells('D5:D7'); ws['D5'] = "JUMLAH PAGU\n(Rp)"
    ws.merge_cells('E5:L5'); ws['E5'] = "SUDAH PENGADAAN"
    ws.merge_cells('M5:M7'); ws['M5'] = "PAKET"
    ws.merge_cells('N5:N7'); ws['N5'] = "Rp"
    ws.merge_cells('E6:F6'); ws['E6'] = "PEMILIHAN/\nPELAKSANAAN"
    ws.merge_cells('G6:H6'); ws['G6'] = "HASIL PEMILIHAN"
    ws.merge_cells('I6:J6'); ws['I6'] = "KONTRAK"
    ws.merge_cells('K6:L6'); ws['K6'] = "SERAH TERIMA"
    ws.row_dimensions[6].height = 30

    headers_row7 = ['PAKET', 'Rp', 'PAKET', 'Rp', 'PAKET', 'Rp', 'PAKET', 'Rp']
    for col_idx, val in enumerate(headers_row7, start=5): ws.cell(row=7, column=col_idx, value=val)

    numbers = ['1','2','3','4','5','6','7','8','9','10','11','12','13=3-5','14=4-6']
    for col_idx, val in enumerate(numbers, start=1): ws.cell(row=8, column=col_idx, value=val)

    for r in range(4, 9):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font_italic if r == 8 else header_font
            cell.fill, cell.alignment, cell.border = header_fill, align_center, border_thin

    if df_rekap.empty: return

    # Isi Data
    row_num = 9
    for idx, row in df_rekap.iterrows():
        is_total = row['Satuan Kerja'] == 'TOTAL KESELURUHAN'
        data_row = [
            "" if is_total else row_num - 8, row['Satuan Kerja'],
            row['Data Laporan_Paket'] if row['Data Laporan_Paket'] else "-", row['Data Laporan_Pagu'] if row['Data Laporan_Pagu'] else "-",
            row['Pemilihan_Paket'] if row['Pemilihan_Paket'] else "-", row['Pemilihan_Pagu'] if row['Pemilihan_Pagu'] else "-",
            row['Hasil_Paket'] if row['Hasil_Paket'] else "-", row['Hasil_Pagu'] if row['Hasil_Pagu'] else "-",
            row['Kontrak_Paket'] if row['Kontrak_Paket'] else "-", row['Kontrak_Pagu'] if row['Kontrak_Pagu'] else "-",
            row['SerahTerima_Paket'] if row['SerahTerima_Paket'] else "-", row['SerahTerima_Pagu'] if row['SerahTerima_Pagu'] else "-",
            row['Belum_Paket'] if row['Belum_Paket'] else "-", row['Belum_Pagu'] if row['Belum_Pagu'] else "-"
        ]
        
        for col_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if is_total:
                cell.font, cell.fill = header_font, header_fill
            else:
                cell.font = font_normal
                cell.fill = PatternFill('solid', start_color='FFFFFF') if row_num % 2 != 0 else PatternFill('solid', start_color='DCE6F1')
            
            cell.border = border_thin
            if col_idx == 2: cell.alignment = align_center if is_total else align_left
            elif col_idx in [4, 6, 8, 10, 12, 14] and isinstance(val, (int, float)):
                cell.number_format, cell.alignment = '#,##0', align_right
            else:
                cell.alignment = align_center
                if isinstance(val, (int, float)): cell.number_format = '#,##0'
                
        ws.row_dimensions[row_num].height = 45
        row_num += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for col in ['C', 'E', 'G', 'I', 'K', 'M']: ws.column_dimensions[col].width = 8
    for col in ['D', 'F', 'H', 'J', 'L', 'N']: ws.column_dimensions[col].width = 16


# =======================================================================================
# PROSES UTAMA
# =======================================================================================
def process_tepra(tahun):
    input_json = os.path.join(BASE_DIR, 'data', str(tahun), f'rekap_pengadaan_{tahun}.json')
    if not os.path.exists(input_json): return
    try:
        with open(input_json, 'r', encoding='utf-8') as f: data = json.load(f)
    except: return
    
    df = pd.DataFrame(data)
    if df.empty: return
    df['Nilai Pagu RUP'] = df['Nilai Pagu RUP'].apply(safe_float)
    df['Nilai Hasil Pemilihan'] = df['Nilai Hasil Pemilihan'].apply(safe_float)

    # 1. TARIK DATA KATEGORI 0 (0 Juta - 50 Juta) --> Batas bawah -1 agar 0 terbawa
    df_k0 = get_data_rekap(df, -1, 50000000)
    
    # 2. TARIK DATA KATEGORI 1 (50 Juta - 200 Juta)
    df_k1 = get_data_rekap(df, 50000000, 200000000)
    
    # 3. TARIK DATA KATEGORI 2 (200 Juta - 2,5 Miliar)
    df_k2 = get_data_rekap(df, 200000000, 2500000000)
    
    # 4. TARIK DATA KATEGORI 3 (2,5 Miliar - 50 Miliar)
    df_k3 = get_data_rekap(df, 2500000000, 50000000000)

    # EXPORT JSON UNTUK WEB
    out_dir_data = os.path.join(BASE_DIR, 'data', str(tahun))
    os.makedirs(out_dir_data, exist_ok=True)
    
    if not df_k0.empty:
        with open(os.path.join(out_dir_data, f'rekap_tepra_0_50m_{tahun}.json'), "w", encoding="utf-8") as f: json.dump(df_k0.to_dict(orient='records'), f, ensure_ascii=False, indent=2)
    if not df_k1.empty:
        with open(os.path.join(out_dir_data, f'rekap_tepra_50_200m_{tahun}.json'), "w", encoding="utf-8") as f: json.dump(df_k1.to_dict(orient='records'), f, ensure_ascii=False, indent=2)
    if not df_k2.empty:
        with open(os.path.join(out_dir_data, f'rekap_tepra_200m_2v5m_{tahun}.json'), "w", encoding="utf-8") as f: json.dump(df_k2.to_dict(orient='records'), f, ensure_ascii=False, indent=2)
    if not df_k3.empty:
        with open(os.path.join(out_dir_data, f'rekap_tepra_2v5m_50m_{tahun}.json'), "w", encoding="utf-8") as f: json.dump(df_k3.to_dict(orient='records'), f, ensure_ascii=False, indent=2)

    # EXPORT KE EXCEL
    out_dir_excel = os.path.join(BASE_DIR, 'output', 'tepra', str(tahun))
    os.makedirs(out_dir_excel, exist_ok=True)
    tgl_gen = datetime.now().strftime("%Y-%m-%d")
    out_excel = os.path.join(out_dir_excel, f'Tepra Kobar_Progres PBJ tahun {tahun} ({tgl_gen}).xlsx')

    wb = Workbook()
    
    # CETAK SHEET 0 (0 - 50 Juta) -> DI POSISI PALING KIRI (Index Active)
    ws0 = wb.active
    ws0.title = "Non strategis 0 sd 50 jt"
    style_excel_sheet_0_50(ws0, df_k0, "PROSES PENGADAAN BARANG DAN JASA PAKET NON STRATEGIS (0 S/D 50 JUTA)")

    # CETAK SHEET 1 (50 Juta - 200 Juta)
    ws1 = wb.create_sheet(title="Non strategis 50jt sd 200jt")
    style_excel_sheet(ws1, df_k1, "PROSES PENGADAAN BARANG DAN JASA PAKET NON STRATEGIS (>50 JUTA - ≤200 JUTA)")

    # CETAK SHEET 2 (200 Juta - 2,5 M)
    ws2 = wb.create_sheet(title="Strategis > 200jt sd 2,5m")
    style_excel_sheet(ws2, df_k2, "PROSES PENGADAAN BARANG DAN JASA PAKET STRATEGIS (> Rp200 JUTA S/D 2,5 M)")

    # CETAK SHEET 3 (2,5 M - 50 M)
    ws3 = wb.create_sheet(title="Strategis > 2,5m sd 50m")
    style_excel_sheet(ws3, df_k3, "PROSES PENGADAAN BARANG DAN JASA PAKET STRATEGIS (≥Rp2,5M sd 50M)")

    wb.save(out_excel)

    # --- JALANKAN FUNGSI ARSIP ---
    kelola_arsip_bulanan(out_dir_excel, tahun)
    update_daftar_arsip_json(out_dir_excel)
    # -----------------------------
    
    print(f"TEPRA {tahun} Sukses (4 Sheet Lengkap) -> {out_excel}")

if __name__ == '__main__':
    process_tepra(tahun_n)