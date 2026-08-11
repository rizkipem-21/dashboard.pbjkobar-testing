# ======================================================
# FASE 2: TRANSFORM & LOAD (GENERATE PAKET PENGADAAN)
# ======================================================

import os
import json
import re
import shutil
import warnings
import pandas as pd
import sys
import subprocess
import time
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import config_rahasia

warnings.filterwarnings('ignore')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

tahun_n      = datetime.now().year       
tahun_n1     = tahun_n - 1               
tahun_n2     = tahun_n - 2               
daftar_tahun = [tahun_n, tahun_n1, tahun_n2] 

# MENGGUNAKAN LOG TUNGGAL (Sama dengan script download)
LOG_FILE = os.path.join(BASE_DIR, 'tools', 'log_pengadaan.txt')
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

def log_print(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    print(msg, **kwargs)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

def get_waktu_indonesia():
    tz_wib = timezone(timedelta(hours=7))
    sekarang = datetime.now(tz_wib)
    bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
    return f"{sekarang.day} {bulan_indo[sekarang.month]} {sekarang.year} | {sekarang.strftime('%H.%M')} WIB"

def sync_to_github():
    log_print("\n==================================================")
    log_print("MENGIRIM DATA PENGADAAN KE GITHUB DARI PYTHON...")
    log_print("==================================================")

    waktu_sekarang = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        subprocess.run(["git", "config", "user.name", "rizkipem-21"], cwd=BASE_DIR)
        subprocess.run(["git", "config", "user.email", "rizki.pem@gmail.com"], cwd=BASE_DIR)
        subprocess.run(["git", "add", "."], capture_output=True, text=True, cwd=BASE_DIR)
        
        commit_msg = f"Auto update Pengadaan {waktu_sekarang}"
        subprocess.run(["git", "commit", "-m", commit_msg], capture_output=True, text=True, cwd=BASE_DIR)
        
        res_push = subprocess.run(["git", "push"], capture_output=True, text=True, cwd=BASE_DIR)
        if res_push.returncode == 0:
            log_print("✅ Push ke GitHub BERHASIL")
            return True, "✅ Push ke GitHub BERHASIL"
        else:
            error_git = res_push.stderr.strip()
            log_print(f"❌ Push ke GitHub GAGAL: {error_git}")
            return False, f"❌ Push ke GitHub GAGAL:\n`{error_git}`"
    except Exception as e:
        log_print(f"❌ Terjadi kesalahan pada eksekusi Git: {str(e)}")
        return False, f"❌ Terjadi kesalahan pada eksekusi Git:\n`{str(e)}`"

def format_tgl(val):
    if not val or (not isinstance(val, str) and pd.isna(val)): return ""
    try:
        parts = str(val).strip()[:10].split('-')
        if len(parts) == 3: return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return ""
    except: return ""

def get_file_path(data_dir, base_name, tahun):
    v1_path = os.path.join(data_dir, f"v1_{base_name}_{tahun}.json")
    if os.path.exists(v1_path): return v1_path
    return os.path.join(data_dir, f"Legacy_{base_name}_{tahun}.json")

def load_json(path):
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
            
            # Jika data adalah list langsung
            if isinstance(data, list): 
                return pd.json_normalize(data) if data else pd.DataFrame()
                
            # Jika data adalah dictionary (respons API standar)
            if isinstance(data, dict):
                for k in ['data', 'items', 'results']:
                    if k in data and isinstance(data[k], list): 
                        return pd.json_normalize(data[k])
                
                return pd.DataFrame() 
                
    except Exception as e:
        return pd.DataFrame()

def kelola_arsip_bulanan(folder_path, tahun):
    if not os.path.exists(folder_path): return
    
    # 1. Buat folder brankas lokal yang tidak akan dibaca GitHub
    folder_arsip_lokal = os.path.join(BASE_DIR, 'arsip_lokal', 'pengadaan', str(tahun))
    os.makedirs(folder_arsip_lokal, exist_ok=True)
    
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    arsip_bulanan = {}
    
    for f in file_excel:
        match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", f)
        if match:
            bulan = match.group(1)[:7]
            if bulan not in arsip_bulanan: arsip_bulanan[bulan] = []
            arsip_bulanan[bulan].append((match.group(1), f))
            
    for bulan, list_file in arsip_bulanan.items():
        # Urutkan dari tanggal paling lama ke paling baru
        list_file.sort(key=lambda x: x[0])
        
        # 2. Pindahkan semua file harian (kecuali yang terbaru) ke folder brankas
        for tgl, nama_file in list_file[:-1]:
            path_sumber = os.path.join(folder_path, nama_file)
            path_tujuan = os.path.join(folder_arsip_lokal, nama_file)
            try: 
                shutil.move(path_sumber, path_tujuan)
                log_print(f"📦 Arsip harian diamankan ke lokal: {nama_file}")
            except Exception as e: 
                log_print(f"⚠️ Gagal memindah arsip {nama_file} (Mungkin file terbuka di Excel): {str(e)}")

def update_daftar_arsip_json(folder_path):
    """Membaca sisa file Excel di folder output dan memperbarui daftar_arsip.json"""
    if not os.path.exists(folder_path): return
    
    # Ambil semua file excel, urutkan dari yang terbaru (Z-A)
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    file_excel.sort(reverse=True) 
    
    # Buat format JSON (sesuaikan dengan format yang dibaca website Anda)
    arsip_list = [{"nama_file": f} for f in file_excel]
    
    # Simpan ke daftar_arsip.json
    path_json = os.path.join(folder_path, 'daftar_arsip.json')
    try:
        with open(path_json, 'w', encoding='utf-8') as f:
            json.dump(arsip_list, f, indent=4)
        log_print(f"📝 Daftar arsip JSON diperbarui: {path_json}")
    except Exception as e:
        log_print(f"⚠️ Gagal memperbarui daftar_arsip.json: {str(e)}")                

def get_kategori_status(sumber, s):
    sumber, s = str(sumber).lower(), str(s).lower()
    if "sumber 1" in sumber: return "Belum Proses"
    if "sumber 2" in sumber or "sumber 5" in sumber: return "Sudah Selesai" if "bapbast" in s else "Sedang Berjalan"
    if "sumber 3" in sumber or "sumber 4" in sumber: return "Sudah Selesai" if "paket selesai" in s or "selesai" in s else "Sedang Berjalan"
    if "sumber 6" in sumber: return "Sudah Selesai" if "payment_outside" in s or "completed" in s else "Sedang Berjalan"
    if "sumber 7" in sumber: return "Sudah Selesai" if "paket selesai" in s or "selesai" in s else "Sedang Berjalan"
    if "bapbast" in s or "payment" in s or "completed" in s: return "Sudah Selesai"
    if "pengumuman rup" in s or s == "" or s == "-": return "Belum Proses"
    return "Sedang Berjalan"

def process_tahun(tahun):
    data_dir = os.path.join(BASE_DIR, 'data', str(tahun))
    output_json = os.path.join(data_dir, f'rekap_pengadaan_{tahun}.json')
    
    if tahun == tahun_n2 and os.path.exists(output_json):
        log_print(f"\n[SKIP] Tahun {tahun} sudah final -> Lewati generate")
        return None

    log_print(f'\n{"="*55}\n   GENERATE DATA TAHUN {tahun}\n{"="*55}')
    def p(nama): return get_file_path(data_dir, nama, tahun)
    
    df1     = load_json(p('rup_paket-penyedia-terumumkan'))
    df1_2   = load_json(p('rup_paket-swakelola-terumumkan'))
    df1_3   = load_json(p('rup_paket-penyedia'))
    df1_4   = load_json(p('rup_paket-swakelola'))
    df1_5   = load_json(p('rup_history-kaji-ulang'))
    
    df2     = load_json(p('tender_non-tender-pengumuman'))
    df2_1   = load_json(p('tender_non-tender-selesai'))
    df2_2   = load_json(p('tender_non-tender-ekontrak-sppbj'))
    df2_3   = load_json(p('tender_non-tender-ekontrak-kontrak'))
    df2_4   = load_json(p('tender_non-tender-ekontrak-spmkspp'))
    df2_5   = load_json(p('tender_non-tender-ekontrak-bapbast'))
    df3     = load_json(p('tender_pencatatan-non-tender'))
    df3_1   = load_json(p('tender_pencatatan-non-tender-realisasi'))
    df4     = load_json(p('tender_pencatatan-swakelola'))
    df5     = load_json(p('tender_pengumuman'))
    df5_1   = load_json(p('tender_tender-selesai'))
    df5_1_1 = load_json(p('tender_tender-selesai-nilai'))
    df5_2   = load_json(p('tender_tender-ekontrak-sppbj'))
    df5_3   = load_json(p('tender_tender-ekontrak-kontrak'))
    df5_4   = load_json(p('tender_tender-ekontrak-spmkspp'))
    df5_5   = load_json(p('tender_tender-ekontrak-bapbast'))
    df6     = load_json(p('ekatalog_paket-e-purchasing'))
    df7     = load_json(p('ekatalog-archive_paket-e-purchasing'))
    df7_1   = load_json(p('ekatalog-archive_instansi-satker'))

    if not df2.empty and 'status_nontender' in df2.columns: df2 = df2[df2['status_nontender'] != 'Gagal/Batal']
    if not df5.empty and 'status_tender' in df5.columns: df5 = df5[df5['status_tender'] != 'Gagal/Batal']
    if not df3.empty and 'status_nontender_pct_ket' in df3.columns: df3 = df3[df3['status_nontender_pct_ket'].astype(str).str.strip() != 'Paket Dibatalkan']
    if not df6.empty and 'status' in df6.columns: df6 = df6[~df6['status'].isin(['CANCELLED_ON_NEGOTIATION', 'CANCELLED_ON_REVIEW', 'CANCELLED'])]

    def get_set(df, col): return set(df[col].astype(str).str.split(';').explode().str.strip()) if not df.empty and col in df.columns else set()

    set_selesai   = get_set(df2_1, 'kd_nontender')
    set_sppbj     = get_set(df2_2, 'kd_nontender')
    set_kontrak   = get_set(df2_3, 'kd_nontender')
    set_spmkspp   = get_set(df2_4, 'kd_nontender')
    set_bapbast   = get_set(df2_5, 'kd_nontender')

    set_t_selesai = get_set(df5_1, 'kd_tender')
    set_t_sppbj   = get_set(df5_2, 'kd_tender')
    set_t_kontrak = get_set(df5_3, 'kd_tender')
    set_t_spmkspp = get_set(df5_4, 'kd_tender')
    set_t_bapbast = get_set(df5_5, 'kd_tender')

    def build_multi_kd_map(df, kd_col, val_col):
        m = {}
        if not df.empty and val_col in df.columns:
            for _, r in df.iterrows():
                for k in str(r.get(kd_col, '')).split(';'):
                    if k.strip(): m[k.strip()] = r.get(val_col)
        return m

    map_nt_kontrak = build_multi_kd_map(df2_1, 'kd_nontender', 'nilai_negosiasi')
    map_nt_pdn     = build_multi_kd_map(df2_1, 'kd_nontender', 'nilai_pdn_kontrak')
    map_nt_umk     = build_multi_kd_map(df2_1, 'kd_nontender', 'nilai_umk_kontrak')
    map_t_kontrak  = build_multi_kd_map(df5_1_1, 'kd_tender', 'nilai_negosiasi')
    map_t_pdn      = build_multi_kd_map(df5_3, 'kd_tender', 'nilai_pdn_kontrak')
    map_t_umk      = build_multi_kd_map(df5_3, 'kd_tender', 'nilai_umk_kontrak')
    
    map_nt_tgl_kontrak = build_multi_kd_map(df2_3, 'kd_nontender', 'tgl_kontrak')
    map_t_tgl_kontrak  = build_multi_kd_map(df5_3, 'kd_tender', 'tgl_kontrak')
    map_nt_penyedia    = build_multi_kd_map(df2_1, 'kd_nontender', 'nama_penyedia')
    map_t_penyedia     = build_multi_kd_map(df5_1_1, 'kd_tender', 'nama_penyedia')

    path_kamus = os.path.join(BASE_DIR, 'data_master', 'kamus_penyedia.json')
    map_offline_penyedia = {}
    if os.path.exists(path_kamus):
        try:
            with open(path_kamus, 'r', encoding='utf-8') as f:
                kamus_list = json.load(f)
                if isinstance(kamus_list, list):
                   for item in kamus_list:
                       nama = item.get('nama_penyedia', "")
                       if item.get('kode_penyedia'): map_offline_penyedia[str(item['kode_penyedia'])] = nama
                       if item.get('kd_penyedia'): map_offline_penyedia[str(item['kd_penyedia'])] = nama
        except: pass

    # =========================================================================
    # BLOK BARU: PEMBERSIHAN DATA & PEMETAAN KAJI ULANG 
    # =========================================================================
    status_terakhir = {} # Menyimpan riwayat terakhir: {rup_id: {'tgl': '...', 'is_batal': True/False}}
    kaji_ulang_dict = {}

    if not df1_5.empty and 'kd_rup_lama' in df1_5.columns and 'kd_rup_baru' in df1_5.columns:
        for _, r in df1_5.iterrows():
            try:
                lama = int(float(str(r['kd_rup_lama']).strip()))
                baru = int(float(str(r['kd_rup_baru']).strip()))
                alasan = str(r.get('alasan_kajiulang', '')).lower()
                jenis_rev = str(r.get('jenis_revisi', '')).lower()
                tgl = str(r.get('tgl_kaji_ulang', ''))
                
                # Cek Indikator Batal dan Aktif (Saringan Kata Kematian & Hakim Waktu)
                kematian = ['batal', 'pembatalan', 'hapus']
                is_batal = any(k in alasan for k in kematian) or any(k in jenis_rev for k in kematian)
                
                # Pengaktifan HANYA SAH jika tidak ada unsur kematian di atas
                is_aktif = ('aktif' in alasan or 'pengaktifan' in jenis_rev) and not is_batal
                
                # Fungsi internal pencatat sejarah
                def catat_sejarah(k_id, batal_flag, tgl_str):
                    if k_id not in status_terakhir:
                        status_terakhir[k_id] = {'tgl': tgl_str, 'is_batal': batal_flag}
                    else:
                        # Update status HANYA JIKA tanggal kaji ulangnya lebih baru/sama
                        if tgl_str >= status_terakhir[k_id]['tgl']:
                            status_terakhir[k_id] = {'tgl': tgl_str, 'is_batal': batal_flag}

                # Tentukan nasib paket berdasarkan waktu
                if is_batal:
                    catat_sejarah(lama, True, tgl)
                    catat_sejarah(baru, True, tgl)
                elif is_aktif:
                    catat_sejarah(lama, False, tgl)
                    catat_sejarah(baru, False, tgl)
                
                # Masukkan ke kamus jika ini revisi normal (bukan batal)
                if lama != baru and not is_batal:
                    kaji_ulang_dict[lama] = baru
            except: pass

    # Eksekusi Blacklist HANYA untuk paket yang nasib TERAKHIRNYA adalah batal
    blacklist_rup = {k for k, v in status_terakhir.items() if v['is_batal']}

    path_manual_konsol = os.path.join(BASE_DIR, 'data_master', 'manual_konsolidasi.json')
    if os.path.exists(path_manual_konsol):
        try:
            with open(path_manual_konsol, 'r', encoding='utf-8') as f:
                manual_data = json.load(f)
                if isinstance(manual_data, list):
                    for item in manual_data:
                        if item.get("status_aktif", False) and str(item.get("tahun_anggaran")) == str(tahun):
                            try:
                                lama = int(float(item["kode_konsol_lama"]))
                                baru = int(float(item["kode_konsol_baru"]))
                                kaji_ulang_dict[lama] = baru
                            except: pass
        except: pass

    def is_related(p1, p2):
        """Sensor: Mengecek relasi maju-mundur antara dua kode RUP"""
        try:
            a, b = int(p1), int(p2)
            if a == b: return True
            # Cek a -> b
            curr, visited = a, set()
            while curr in kaji_ulang_dict and curr not in visited:
                visited.add(curr)
                curr = kaji_ulang_dict[curr]
                if curr == b: return True
            # Cek b -> a
            curr, visited = b, set()
            while curr in kaji_ulang_dict and curr not in visited:
                visited.add(curr)
                curr = kaji_ulang_dict[curr]
                if curr == a: return True
            return False
        except: return False

    def is_related_list(l_list, r_list):
        """Memastikan Kiri memiliki ikatan sejarah dengan Kanan secara sejajar"""
        if l_list == r_list: return True
        for i in range(len(l_list)):
            if not is_related(l_list[i], r_list[i]): return False
        return True

    def bersihkan_dan_ambil_kanan(raw_rup_str):
        if pd.isna(raw_rup_str) or not str(raw_rup_str).strip(): return ""
        parts = [p.strip() for p in str(raw_rup_str).split(';') if p.strip()]
        
        if len(parts) <= 1:
            return ";".join(parts)
            
        # =================================================================
        # TAHAP 1: PEMETAAN SILSILAH (Melihat Gambaran Besar)
        # =================================================================
        families = []
        for p in parts:
            found_family = False
            for fam in families:
                # Jika p punya riwayat dengan salah satu anggota keluarga ini
                if any(is_related(p, member) for member in fam):
                    fam.append(p)
                    found_family = True
                    break
            if not found_family:
                families.append([p]) # Buat keluarga baru jika mandiri
                
        # =================================================================
        # TAHAP 2: KLASIFIKASI & EKSEKUSI
        # =================================================================
        
        # SKENARIO 1: 1 HISTORY PANJANG (Baik Genap/Ganjil)
        if len(families) == 1:
            return parts[-1] # Langsung comot paling kanan
            
        # SKENARIO 2: PASANGAN IDENTIK (Acak / Cermin / Berurutan)
        if len(parts) % 2 == 0:
            mid = len(parts) // 2
            kiri = parts[:mid]
            kanan = parts[mid:]
            
            # Cek silang apakah Kiri dan Kanan punya pasangan tepat 1-1
            is_pasangan_identik = True
            kanan_terpakai = set()
            
            for k in kiri:
                punya_pasangan = False
                for idx_kn, kn in enumerate(kanan):
                    if idx_kn not in kanan_terpakai and is_related(k, kn):
                        punya_pasangan = True
                        kanan_terpakai.add(idx_kn)
                        break
                if not punya_pasangan:
                    is_pasangan_identik = False
                    break
                    
            if is_pasangan_identik:
                return ";".join(kanan) # Belah 2, Ambil mutlak Kanan
                
        # SKENARIO 3: LONCATAN / JOMBLO / CAMPURAN (A B C)
        # Eksekusi: Ambil 1 kode paling kanan dari setiap keluarga
        rightmost_set = set(fam[-1] for fam in families)
        final_parts = []
        
        # Mempertahankan urutan cetak aslinya
        for p in parts:
            if p in rightmost_set:
                final_parts.append(p)
                rightmost_set.remove(p) # Kunci agar tidak duplikat
                
        return ";".join(final_parts)

    def cari_rup_baru(kode_rup_bersih):
        """Melacak RUP sampai ke ujung sejarahnya (Looping)"""
        if not kode_rup_bersih: return ""
        parts = kode_rup_bersih.split(';')
        results = []
        for p in parts:
            if p.isdigit():
                curr = int(p)
                if curr in blacklist_rup: continue
                
                awal = curr
                visited = set()
                
                # Telusuri terus ke depan sampai ujung rantai
                while curr in kaji_ulang_dict and curr not in visited:
                    visited.add(curr)
                    next_rup = kaji_ulang_dict[curr]
                    # Berhenti jika ujungnya ternyata paket mati
                    if next_rup in blacklist_rup:
                        curr = None
                        break
                    curr = next_rup
                
                if curr is not None and curr != awal:
                    results.append(str(curr))
        return ";".join(list(dict.fromkeys(results)))
    # =========================================================================

    def split_kd_list(x): return [int(i.strip()) for i in str(x).split(';') if i.strip().isdigit()]

    def standardize_kd_rup(df, col):
        if df.empty or col not in df.columns: return df
        df[col+'_raw'] = df[col]
        df[col+'_list'] = df[col].apply(split_kd_list)
        df[col] = df[col].apply(lambda x: split_kd_list(x)[0] if len(split_kd_list(x))>0 else None)
        return df

    for d, c in [(df1,'kd_rup'), (df1_2,'kd_rup'), (df1_3,'kd_rup'), (df1_4,'kd_rup'), (df2,'kd_rup'), (df3,'kd_rup'), (df4,'kd_rup'), (df5,'kd_rup'), (df6,'rup_code'), (df7,'kd_rup')]:
        standardize_kd_rup(d, c)

    map_pagu_s1   = df1.drop_duplicates(subset=['kd_rup']).set_index('kd_rup')['pagu'] if not df1.empty and 'kd_rup' in df1.columns else {}
    map_pagu_s1_2 = df1_2.drop_duplicates(subset=['kd_rup']).set_index('kd_rup')['pagu'] if not df1_2.empty and 'kd_rup' in df1_2.columns else {}
    map_pagu_s1_3 = df1_3.drop_duplicates(subset=['kd_rup']).set_index('kd_rup')['pagu'] if not df1_3.empty and 'kd_rup' in df1_3.columns else {}
    map_pagu_s1_4 = df1_4.drop_duplicates(subset=['kd_rup']).set_index('kd_rup')['pagu'] if not df1_4.empty and 'kd_rup' in df1_4.columns else {}

    def get_pagu_multi(kd_list, tipe='s1'):
        if not kd_list: return None
        total, added = 0, set() 
        for k in kd_list:
            try: k_int = int(float(str(k).strip()))
            except: continue
            
            if k_int in added: continue
            added.add(k_int)
            
            if tipe == 's1':
                val = map_pagu_s1.get(k_int, map_pagu_s1_3.get(k_int, 0))
            else:
                val = map_pagu_s1_2.get(k_int, map_pagu_s1_4.get(k_int, 0))
                
            if (pd.isna(val) or float(val) == 0) and k_int in kaji_ulang_dict:
                k_baru = kaji_ulang_dict[k_int]
                if tipe == 's1':
                    val = map_pagu_s1.get(k_baru, map_pagu_s1_3.get(k_baru, 0))
                else:
                    val = map_pagu_s1_2.get(k_baru, map_pagu_s1_4.get(k_baru, 0))
                    
            total += float(val) if not pd.isna(val) else 0
        return total if total != 0 else None

    def build_df_map(df_source):
        if not df_source.empty:
            df_clean = df_source.dropna(subset=['kd_rup']).drop_duplicates(subset=['kd_rup']).copy()
            df_clean['kd_rup'] = df_clean['kd_rup'].apply(lambda x: int(float(str(x).strip())))
            return df_clean.set_index('kd_rup')
        return pd.DataFrame()

    df1_map, df1_2_map, df1_3_map, df1_4_map = build_df_map(df1), build_df_map(df1_2), build_df_map(df1_3), build_df_map(df1_4)
    map_status_konsol = {}
    if not df1_3.empty and 'kd_rup' in df1_3.columns and 'status_konsolidasi' in df1_3.columns:
        map_status_konsol = {int(float(str(k).strip())): v for k, v in zip(df1_3['kd_rup'], df1_3['status_konsolidasi']) if pd.notna(k) and str(k).strip().isdigit()}

    def get_s1(kd, col, tipe='s1'):
        try:
            if pd.isna(kd): return None
            kd_match = int(float(str(kd).strip()))
            if tipe == 's1':
                if not df1_map.empty and kd_match in df1_map.index and not pd.isna(df1_map.loc[kd_match, col]): return df1_map.loc[kd_match, col]
                if not df1_3_map.empty and kd_match in df1_3_map.index and not pd.isna(df1_3_map.loc[kd_match, col]): return df1_3_map.loc[kd_match, col]
            elif tipe == 's1_2':
                if not df1_2_map.empty and kd_match in df1_2_map.index and not pd.isna(df1_2_map.loc[kd_match, col]): return df1_2_map.loc[kd_match, col]
                if not df1_4_map.empty and kd_match in df1_4_map.index and not pd.isna(df1_4_map.loc[kd_match, col]): return df1_4_map.loc[kd_match, col]
            return None
        except: return None

    executed_rups = set()
    for df_trans, col_name in [(df2,'kd_rup_raw'), (df3,'kd_rup_raw'), (df4,'kd_rup_raw'), (df5,'kd_rup_raw'), (df6,'rup_code_raw'), (df7,'kd_rup_raw')]:
        if not df_trans.empty and col_name in df_trans.columns:
            for raw_k in df_trans[col_name]:
                bersih = bersihkan_dan_ambil_kanan(raw_k)
                for k_str in bersih.split(';'):
                    if k_str.strip().isdigit():
                        curr = int(k_str.strip())
                        executed_rups.add(curr)
                        # Lacak sampai ujung riwayat agar RUP Baru tidak lolos ke Pengumuman RUP
                        visited = set()
                        while curr in kaji_ulang_dict and curr not in visited:
                            visited.add(curr)
                            curr = kaji_ulang_dict[curr]
                            executed_rups.add(curr)

    data_s2=[]
    for _, r in df2.iterrows():
        raw_rup = r.get('kd_rup_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('kd_rup')
        kd_nt_list = [i.strip() for i in str(r.get('kd_nontender', '')).split(';')] if pd.notna(r.get('kd_nontender')) else []
        status_awal = r.get('status_nontender')
        
        if pd.notna(status_awal) and 'berlangsung' in str(status_awal).lower():
            found_in_selesai = any(k in s for k in kd_nt_list for s in [set_bapbast, set_spmkspp, set_kontrak, set_sppbj, set_selesai])
            status = 'Berlangsung' if found_in_selesai else 'Pemilihan Penyedia'
        else:
            status = status_awal
            for k in kd_nt_list:
                if k in set_bapbast: status='BAPBAST'; break
                elif k in set_spmkspp: status='SPMKSPP'; break
                elif k in set_kontrak: status='Kontrak'; break
                elif k in set_sppbj: status='SPPBJ'; break
                elif k in set_selesai: status='Non Tender Selesai'; break

        data_s2.append({
            'Kode Paket': r.get('kd_nontender'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'),
            'Metode Pemilihan': r.get('mtd_pemilihan'), 'Jenis Pengadaan': r.get('jenis_pengadaan'), 'Sumber Dana': r.get('sumber_dana'),
            'PDN': get_s1(kd_lookup, 'status_pdn', 's1'), 'UKM': get_s1(kd_lookup, 'status_ukm', 's1'), 'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1'),
            'Nilai Hasil Pemilihan': next((map_nt_kontrak[k] for k in kd_nt_list if k in map_nt_kontrak), "N/A"), 
            'Tanggal Kontrak': format_tgl(next((map_nt_tgl_kontrak[k] for k in kd_nt_list if k in map_nt_tgl_kontrak), "")),
            'Nama Penyedia': next((map_nt_penyedia[k] for k in kd_nt_list if k in map_nt_penyedia), ""), 'Status': status,
            'Nilai HPS': r.get('hps'), 'Nilai PDN': next((map_nt_pdn[k] for k in kd_nt_list if k in map_nt_pdn), "N/A"), 
            'Nilai UMK': next((map_nt_umk[k] for k in kd_nt_list if k in map_nt_umk), "N/A"),
            'Cara Pengadaan': 'Non Tender', 'Sumber': 'Sumber 2'
        })
    df_s2 = pd.DataFrame(data_s2)

    map_s3_penyedia = {}
    if not df3_1.empty and 'nama_penyedia' in df3_1.columns:
        kd_kolom_3_1 = next((c for c in df3_1.columns if c.lower().strip() in ['kd_nontender_pct', 'kd_nontender', 'kode_paket']), None)
        if kd_kolom_3_1:
            for pkt_id, group in df3_1.groupby(kd_kolom_3_1):
                penyedia_set = set()
                for _, row in group.iterrows():
                    nama = str(row.get('nama_penyedia', '')).strip()
                    npwp = str(row.get('npwp_penyedia', '')).strip()
                    if nama and nama.lower() not in ['', 'nan', 'none', '-']:
                        penyedia_set.add((nama, npwp))
                
                daftar_nama = [nama for nama, npwp in penyedia_set]
                if daftar_nama:
                    map_s3_penyedia[str(pkt_id).strip()] = "; ".join(daftar_nama)

    data_s3=[]
    for _, r in df3.iterrows():
        raw_rup = r.get('kd_rup_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('kd_rup')
        
        data_s3.append({
            'Kode Paket': r.get('kd_nontender_pct'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'),
            'Metode Pemilihan': r.get('mtd_pemilihan'), 'Jenis Pengadaan': r.get('kategori_pengadaan'), 'Sumber Dana': r.get('sumber_dana'),
            'PDN': get_s1(kd_lookup, 'status_pdn', 's1'), 'UKM': get_s1(kd_lookup, 'status_ukm', 's1'), 'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1'),
            'Nilai Hasil Pemilihan': "" if pd.isna(r.get('total_realisasi')) else r.get('total_realisasi'), 'Tanggal Kontrak': format_tgl(r.get('tgl_selesai_paket', '')),
            'Nama Penyedia': map_s3_penyedia.get(str(r.get('kd_nontender_pct')).strip(), ""), 'Status': r.get('status_nontender_pct_ket'), 'Nilai HPS': pd.NA,
            'Nilai PDN': r.get('nilai_pdn_pct'), 'Nilai UMK': r.get('nilai_umk_pct'), 'Cara Pengadaan': 'Pencatatan Non Tender', 'Sumber': 'Sumber 3'
        })
    df_s3 = pd.DataFrame(data_s3)

    data_s4=[]
    swakelola_map = df1_2.set_index('kd_rup')['tipe_swakelola'] if not df1_2.empty else {}
    swakelola_map_fallback = df1_4.set_index('kd_rup')['tipe_swakelola'] if not df1_4.empty else {}
    
    for _, r in df4.iterrows():
        raw_rup = r.get('kd_rup_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('kd_rup')

        jenis = f"Swakelola {int(swakelola_map[kd_lookup])}" if kd_lookup in swakelola_map else (f"Swakelola {int(swakelola_map_fallback[kd_lookup])}" if kd_lookup in swakelola_map_fallback else "N/A")
        
        data_s4.append({
            'Kode Paket': r.get('kd_swakelola_pct'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'),
            'Metode Pemilihan': 'Swakelola', 'Jenis Pengadaan': jenis, 'Sumber Dana': r.get('sumber_dana'),
            'PDN': "PDN" if r.get('nilai_pdn_pct', 0)!=0 else "Tidak", 'UKM': "UKM" if r.get('nilai_umk_pct', 0)!=0 else "Tidak",
            'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1_2'), 'Nilai Hasil Pemilihan': "" if pd.isna(r.get('total_realisasi')) else r.get('total_realisasi'),
            'Tanggal Kontrak': format_tgl(r.get('tgl_selesai_paket', '')), 'Nama Penyedia': "", 'Status': r.get('status_swakelola_pct_ket'),
            'Nilai HPS': pd.NA, 'Nilai PDN': r.get('nilai_pdn_pct'), 'Nilai UMK': r.get('nilai_umk_pct'),
            'Cara Pengadaan': 'Pencatatan Swakelola', 'Sumber': 'Sumber 4'
        })
    df_s4 = pd.DataFrame(data_s4)

    data_s5=[]
    for _, r in df5.iterrows():
        raw_rup = r.get('kd_rup_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('kd_rup')
        kd_t_list = [i.strip() for i in str(r.get('kd_tender', '')).split(';')] if pd.notna(r.get('kd_tender')) else []
        status_awal = r.get('status_tender')
        
        if pd.notna(status_awal) and 'berlangsung' in str(status_awal).lower():
            found_in_selesai = any(k in s for k in kd_t_list for s in [set_t_bapbast, set_t_spmkspp, set_t_kontrak, set_t_sppbj, set_t_selesai])
            status = 'Berlangsung' if found_in_selesai else 'Pemilihan Penyedia'
        else:
            status = status_awal
            for k in kd_t_list:
                if k in set_t_bapbast: status='BAPBAST'; break
                elif k in set_t_spmkspp: status='SPMKSPP'; break
                elif k in set_t_kontrak: status='Kontrak'; break
                elif k in set_t_sppbj: status='SPPBJ'; break
                elif k in set_t_selesai: status='Tender Selesai'; break

        mtd = str(r.get('mtd_pemilihan', '')).strip().lower()
        kat_metode = 'Seleksi' if mtd == 'seleksi' else ('Tender Cepat' if mtd == 'tender cepat' else 'Tender')

        data_s5.append({
            'Kode Paket': r.get('kd_tender'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'),
            'Metode Pemilihan': r.get('mtd_pemilihan'), 'Jenis Pengadaan': r.get('jenis_pengadaan'), 'Sumber Dana': r.get('sumber_dana'),
            'PDN': get_s1(kd_lookup, 'status_pdn', 's1'), 'UKM': get_s1(kd_lookup, 'status_ukm', 's1'), 'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1'),
            'Nilai Hasil Pemilihan': next((map_t_kontrak[k] for k in kd_t_list if k in map_t_kontrak), "N/A"), 
            'Tanggal Kontrak': format_tgl(next((map_t_tgl_kontrak[k] for k in kd_t_list if k in map_t_tgl_kontrak), "")),
            'Nama Penyedia': next((map_t_penyedia[k] for k in kd_t_list if k in map_t_penyedia), ""), 'Status': status,
            'Nilai HPS': r.get('hps'), 'Nilai PDN': next((map_t_pdn[k] for k in kd_t_list if k in map_t_pdn), "N/A"), 
            'Nilai UMK': next((map_t_umk[k] for k in kd_t_list if k in map_t_umk), "N/A"),
            'Cara Pengadaan': kat_metode, 'Sumber': 'Sumber 5'
        })
    df_s5 = pd.DataFrame(data_s5)

    data_s6=[]
    for _, r in df6.iterrows():
        raw_rup = r.get('rup_code_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('rup_code')
        nilai_hasil = r.get('total') if not pd.isna(r.get('total')) else ""
        
        try:
            status_pdn_katalog = get_s1(kd_lookup, 'status_pdn', 's1')
            status_ukm_katalog = get_s1(kd_lookup, 'status_ukm', 's1')
            nilai_pdn_val = nilai_hasil if str(status_pdn_katalog).strip().upper() == 'PDN' else 0
            nilai_umk_val = nilai_hasil if str(status_ukm_katalog).strip().upper() == 'UKM' else 0
        except:
            status_pdn_katalog, status_ukm_katalog, nilai_pdn_val, nilai_umk_val = "N/A", "N/A", "N/A", "N/A"

        kode_p = str(r.get('kode_penyedia', ""))
        nama_p = map_offline_penyedia.get(kode_p, kode_p) if kode_p and kode_p != "None" else ""

        data_s6.append({
            'Kode Paket': r.get('order_id'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('rup_name'),
            'Metode Pemilihan': 'E-Purchasing', 'Jenis Pengadaan': get_s1(kd_lookup, 'jenis_pengadaan', 's1'), 'Sumber Dana': r.get('funding_source'),
            'PDN': status_pdn_katalog, 'UKM': status_ukm_katalog, 'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1'),
            'Nilai Hasil Pemilihan': nilai_hasil, 'Tanggal Kontrak': "", 'Nama Penyedia': nama_p, 'Status': r.get('status'),
            'Nilai HPS': pd.NA, 'Nilai PDN': nilai_pdn_val, 'Nilai UMK': nilai_umk_val,
            'Cara Pengadaan': 'E-Purchasing V6', 'Sumber': 'Sumber 6'
        })
    df_s6 = pd.DataFrame(data_s6)

    data_s7=[]
    map_satker_v5 = {}
    if not df7_1.empty:
        for _, row_s in df7_1.iterrows():
            kd_s = row_s.get('kd_satker')
            if pd.notna(kd_s):
                try: map_satker_v5[str(int(float(kd_s)))] = row_s.get('nama_satker', '')
                except: map_satker_v5[str(kd_s).strip()] = row_s.get('nama_satker', '')

    for _, r in df7.iterrows():
        raw_rup = r.get('kd_rup_raw')
        kode_rup_asli_kotor = bersihkan_dan_ambil_kanan(raw_rup)
        
        # Saringan Konsolidasi: Buang RUP yang batal, sisakan yang hidup
        cleaned_list = [int(i.strip()) for i in kode_rup_asli_kotor.split(';') if i.strip().isdigit() and int(i.strip()) not in blacklist_rup]
        
        # Buang baris paket INI HANYA JIKA semua RUP-nya batal
        if not cleaned_list: continue 
        
        kode_rup_asli = ";".join(map(str, cleaned_list))
        kode_rup_baru = cari_rup_baru(kode_rup_asli)

        kd_lookup = cleaned_list[0] if cleaned_list else r.get('kd_rup')
        nilai_hasil = r.get('total_harga') if not pd.isna(r.get('total_harga')) else ""
        
        try:
            status_pdn_katalog = get_s1(kd_lookup, 'status_pdn', 's1')
            status_ukm_katalog = get_s1(kd_lookup, 'status_ukm', 's1')
            nilai_pdn_val = nilai_hasil if str(status_pdn_katalog).strip().upper() == 'PDN' else 0
            nilai_umk_val = nilai_hasil if str(status_ukm_katalog).strip().upper() == 'UKM' else 0
        except:
            status_pdn_katalog, status_ukm_katalog, nilai_pdn_val, nilai_umk_val = "N/A", "N/A", "N/A", "N/A"

        kode_p = str(r.get('kd_penyedia', ""))
        nama_p = map_offline_penyedia.get(kode_p, kode_p) if kode_p and kode_p != "None" else ""
        sid = r.get('satker_id')
        try: key_id = str(int(float(sid))) if pd.notna(sid) else ""
        except: key_id = str(sid).strip() if pd.notna(sid) else ""
        
        paket_status_str = r.get('paket_status_str')
        status_s7 = r.get('status_paket') if paket_status_str == "Paket Proses" else (r.get('paket_status_str') if paket_status_str == "Paket Selesai" else paket_status_str)

        data_s7.append({
            'Kode Paket': r.get('no_paket'), 'Kode RUP': kode_rup_asli, 'Kode RUP Baru': kode_rup_baru,
            'Satuan Kerja': map_satker_v5.get(key_id, ""),
            'Nama Paket': r.get('nama_paket'), 'Metode Pemilihan': 'E-Purchasing', 'Jenis Pengadaan': get_s1(kd_lookup, 'jenis_pengadaan', 's1'), 'Sumber Dana': r.get('nama_sumber_dana'),
            'PDN': status_pdn_katalog, 'UKM': status_ukm_katalog, 'Nilai Pagu RUP': get_pagu_multi(cleaned_list, 's1'),
            'Nilai Hasil Pemilihan': nilai_hasil, 'Tanggal Kontrak': "", 'Nama Penyedia': nama_p, 
            'Status': status_s7, 'Nilai HPS': pd.NA, 'Nilai PDN': nilai_pdn_val, 'Nilai UMK': nilai_umk_val,
            'Cara Pengadaan': 'E-Purchasing V5', 'Sumber': 'Sumber 7'
        })
    df_s7 = pd.DataFrame(data_s7)

    data_s1_2=[]
    for _, r in df1_2.iterrows():
        kd = r.get('kd_rup')
        try:
            kd_int = int(float(str(kd).strip()))
            if kd_int in blacklist_rup: continue
        except: pass
            
        if kd not in executed_rups:
            jenis = f"Swakelola {int(swakelola_map[kd])}" if kd in swakelola_map else "N/A"
            kode_rup_baru_s1 = cari_rup_baru(str(kd)) if pd.notna(kd) else ""
            data_s1_2.append({
                'Kode Paket': pd.NA, 'Kode RUP': kd, 'Kode RUP Baru': kode_rup_baru_s1,
                'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'), 'Metode Pemilihan': 'Swakelola',
                'Jenis Pengadaan': jenis, 'Sumber Dana': None, 'PDN': None, 'UKM': None, 'Nilai Pagu RUP': r.get('pagu'), 'Nilai Hasil Pemilihan': "",
                'Tanggal Kontrak': "", 'Nama Penyedia': "", 'Status': 'Pengumuman RUP', 'Nilai HPS': pd.NA, 'Nilai PDN': pd.NA, 'Nilai UMK': pd.NA,
                'Cara Pengadaan': 'Swakelola', 'Sumber': 'Sumber 1_2'
            })
    df_s1_2 = pd.DataFrame(data_s1_2)

    data_s1=[]
    for _, r in df1.iterrows():
        kd = r.get('kd_rup')
        try:
            kd_int = int(float(str(kd).strip()))
            if kd_int in blacklist_rup: continue
        except: pass
            
        if kd not in executed_rups:
            kode_rup_baru_s1 = cari_rup_baru(str(kd)) if pd.notna(kd) else ""
            data_s1.append({
                'Kode Paket': pd.NA, 'Kode RUP': kd, 'Kode RUP Baru': kode_rup_baru_s1,
                'Satuan Kerja': r.get('nama_satker'), 'Nama Paket': r.get('nama_paket'), 'Metode Pemilihan': r.get('metode_pengadaan'),
                'Jenis Pengadaan': r.get('jenis_pengadaan'), 'Sumber Dana': None, 'PDN': 'PDN' if r.get('status_pdn')=='PDN' else 'Non-PDN',
                'UKM': 'UKM' if r.get('status_ukm')=='UKM' else 'Non-UKM', 'Nilai Pagu RUP': r.get('pagu'), 'Nilai Hasil Pemilihan': "",
                'Tanggal Kontrak': "", 'Nama Penyedia': "", 'Status': 'Pengumuman RUP', 'Nilai HPS': pd.NA, 'Nilai PDN': pd.NA, 'Nilai UMK': pd.NA,
                'Cara Pengadaan': r.get('metode_pengadaan'), 'Sumber': 'Sumber 1'
            })
    df_s1 = pd.DataFrame(data_s1)

    final_df = pd.concat([df_s2, df_s3, df_s4, df_s1_2, df_s5, df_s6, df_s7, df_s1], ignore_index=True)
    final_df = final_df.drop_duplicates(ignore_index=True)
    final_df = final_df.map(lambda x: re.sub(r'[\x00-\x1F]', '', str(x)) if isinstance(x, str) else x)

    def get_status_konsolidasi(rup_val):
        if pd.isna(rup_val) or not str(rup_val).strip(): return ""
        valid_rups = [k for k in str(rup_val).split(';') if k.strip().isdigit()]
        if len(valid_rups) > 1: return "Konsolidasi"
        elif len(valid_rups) == 1: return map_status_konsol.get(int(valid_rups[0]), "")
        return ""
    final_df['Status Konsolidasi'] = final_df['Kode RUP'].apply(get_status_konsolidasi)

    cols = ['Kode Paket', 'Kode RUP', 'Kode RUP Baru', 'Satuan Kerja', 'Nama Paket', 'Metode Pemilihan', 'Jenis Pengadaan', 'Sumber Dana', 'PDN', 'UKM', 'Status Konsolidasi', 'Nilai Pagu RUP', 'Nilai Hasil Pemilihan', 'Tanggal Kontrak', 'Nama Penyedia', 'Status', 'Nilai HPS', 'Nilai PDN', 'Nilai UMK', 'Cara Pengadaan', 'Sumber']
    final_df = final_df[cols].fillna("")
    final_df['PDN'] = final_df['PDN'].replace("", "N/A")
    final_df['UKM'] = final_df['UKM'].replace("", "N/A")
    final_df['Status'] = final_df['Status'].apply(lambda x: str(x).replace('_', ' ') if pd.notna(x) else x)

    def safe_numeric(val):
        if val in ("N/A", "", None) or pd.isna(val): return 0.0
        try: return float(val)
        except: return 0.0

    def aggregate_text(series):
        valid = [str(x).strip() for x in series if pd.notna(x) and str(x).strip() not in ("N/A", "")]
        return "; ".join(list(dict.fromkeys(valid)))

    def aggregate_raw_status(group, pagu_rup, sum_hasil):
        def get_score(s_text):
            s_low = str(s_text).strip().lower()
            if 'batal' in s_low or 'gagal' in s_low or 'cancel' in s_low: return -1
            elif 'bapbast' in s_low or 'payment' in s_low or 'completed' in s_low or 'paket selesai' in s_low: return 100
            elif 'spmkspp' in s_low: return 60
            elif 'kontrak' in s_low: return 50
            elif 'sppbj' in s_low: return 40
            elif 'selesai' in s_low: return 30
            elif 'pengumuman' in s_low and 'rup' not in s_low: return 20
            elif 'pengumuman rup' in s_low: return 10
            elif s_low == '' or s_low == '-': return 0
            else: return 5

        scored_statuses = []
        all_batal, has_proses_berjalan = True, False
        
        for _, row in group.iterrows():
            raw_s = str(row['Status']).strip()
            score = get_score(raw_s)
            if score != -1: all_batal = False
            if 0 < score < 100 and score != 10: has_proses_berjalan = True
            if raw_s: scored_statuses.append((score, raw_s))
            
        if all_batal or not scored_statuses: return "; ".join(list(dict.fromkeys([s for score, s in scored_statuses])))
        
        valid_scored = [item for item in scored_statuses if item[0] != -1]
        if not valid_scored: return "; ".join(list(dict.fromkeys([s for score, s in scored_statuses])))

        if has_proses_berjalan:
            filtered_berjalan = [item for item in valid_scored if item[0] < 100]
            if filtered_berjalan:
                max_score = max([score for score, s in filtered_berjalan])
                return "; ".join(list(dict.fromkeys([s for score, s in filtered_berjalan if score == max_score])))

        max_score = max([score for score, s in valid_scored])
        hasil_status = "; ".join(list(dict.fromkeys([s for score, s in valid_scored if score == max_score])))
        
        persentase_terpakai = (sum_hasil / pagu_rup) * 100 if pagu_rup > 0 else 100
        if len(group) > 1 and max_score == 100 and persentase_terpakai < 80: return "Sedang Berjalan (Sisa Pagu)"
        return hasil_status

    df_agregasi = final_df.copy()
    for c in ['Nilai HPS', 'Nilai Hasil Pemilihan', 'Nilai PDN', 'Nilai UMK', 'Nilai Pagu RUP']:
        df_agregasi[c] = df_agregasi[c].apply(safe_numeric)

    df_agregasi['Kode RUP_Group'] = df_agregasi['Kode RUP'].apply(lambda x: str(x).split(';')[0].strip() if x else "UNKNOWN")
    grouped = df_agregasi.groupby('Kode RUP_Group')

    rekap_data = []
    for kode_rup, group in grouped:
        first_row = group.iloc[0]
        sum_hps, sum_hasil = group['Nilai HPS'].sum(), group['Nilai Hasil Pemilihan'].sum()
        sum_pdn, sum_umk = group['Nilai PDN'].sum(), group['Nilai UMK'].sum()
        pagu_rup = first_row['Nilai Pagu RUP']

        rekap_data.append({
            'Kode Paket': aggregate_text(group['Kode Paket']), 'Kode RUP': first_row['Kode RUP'], 'Kode RUP Baru': first_row['Kode RUP Baru'],
            'Satuan Kerja': first_row['Satuan Kerja'], 'Nama Paket': first_row['Nama Paket'], 'Metode Pemilihan': first_row['Metode Pemilihan'],
            'Jenis Pengadaan': first_row['Jenis Pengadaan'], 'Sumber Dana': first_row['Sumber Dana'], 'PDN': first_row['PDN'], 'UKM': first_row['UKM'], 'Status Konsolidasi': first_row['Status Konsolidasi'],
            'Nilai Pagu RUP': pagu_rup if pagu_rup != 0 else "", 'Nilai Hasil Pemilihan': sum_hasil if sum_hasil != 0 else "",
            'Tanggal Kontrak': aggregate_text(group['Tanggal Kontrak']), 'Nama Penyedia': aggregate_text(group['Nama Penyedia']),
            'Status': aggregate_raw_status(group, pagu_rup, sum_hasil), 'Nilai HPS': sum_hps if sum_hps != 0 else "",
            'Nilai PDN': sum_pdn if sum_pdn != 0 else "", 'Nilai UMK': sum_umk if sum_umk != 0 else "",
            'Cara Pengadaan': aggregate_text(group['Cara Pengadaan']), 'Sumber': aggregate_text(group['Sumber'])
        })
    df_rekap = pd.DataFrame(rekap_data)

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(final_df.to_dict(orient='records'), f, ensure_ascii=False, indent=2)
    log_print(f"JSON Rekap sukses dibuat: {output_json}")

    kolom_angka_baku = ['Nilai Pagu RUP', 'Nilai Hasil Pemilihan', 'Nilai HPS', 'Nilai PDN', 'Nilai UMK']
    excel_df_detail = final_df.copy()
    for col in kolom_angka_baku:
        if col in excel_df_detail.columns:
            excel_df_detail[col] = excel_df_detail[col].apply(lambda x: safe_numeric(x) if safe_numeric(x) != 0 else "")

    tahun_label = str(df1['tahun_anggaran'].iloc[0]) if not df1.empty and 'tahun_anggaran' in df1.columns else str(tahun)
    nama_file_excel = f'Paket Pengadaan Tahun {tahun_label} ({datetime.now().strftime("%Y-%m-%d")}).xlsx'
    output_dir_excel = os.path.join(BASE_DIR, 'output', 'pengadaan', str(tahun))
    os.makedirs(output_dir_excel, exist_ok=True)
    output_excel_path = os.path.join(output_dir_excel, nama_file_excel)

    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_rekap.to_excel(writer, index=False, sheet_name='1. Rekap per RUP')
        excel_df_detail.to_excel(writer, index=False, sheet_name='2. Detail per Paket')

    wb = load_workbook(output_excel_path)
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=10)
    fill_putih = PatternFill('solid', start_color='FFFFFF')
    fill_biru_muda = PatternFill('solid', start_color='DCE6F1')
    border_thin = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

    def style_sheet(ws_name, df_ref, dict_lebar, list_kolom_angka):
        ws = wb[ws_name]
        for i, col in enumerate(df_ref.columns, start=1): ws.column_dimensions[get_column_letter(i)].width = dict_lebar.get(col, 15)
        for cell in ws[1]: cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, border_thin
        ws.row_dimensions[1].height = 32
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            fill = fill_putih if row_idx % 2 == 0 else fill_biru_muda
            for cell in row:
                col_name = df_ref.columns[cell.column - 1]
                cell.font, cell.fill, cell.border = data_font, fill, border_thin
                if col_name in list_kolom_angka and str(cell.value).replace('.', '', 1).isdigit():
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else: cell.alignment = Alignment(vertical='center', wrap_text=False)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    lebar_baku = {'Kode Paket': 25, 'Kode RUP': 18, 'Kode RUP Baru': 18, 'Satuan Kerja': 38, 'Nama Paket': 50, 'Metode Pemilihan': 22, 'Jenis Pengadaan': 32, 'Sumber Dana': 14, 'PDN': 10, 'UKM': 10, 'Status Konsolidasi': 18, 'Nilai Pagu RUP': 20, 'Nilai Hasil Pemilihan': 20, 'Tanggal Kontrak': 25, 'Nama Penyedia': 40, 'Status': 28, 'Nilai HPS': 20, 'Nilai PDN': 18, 'Nilai UMK': 18, 'Cara Pengadaan': 25, 'Sumber': 15}
    style_sheet('1. Rekap per RUP', df_rekap, lebar_baku, kolom_angka_baku)
    style_sheet('2. Detail per Paket', excel_df_detail, lebar_baku, kolom_angka_baku)
    wb.save(output_excel_path)
    
    shutil.copy2(output_excel_path, os.path.join(data_dir, f'master_pengadaan_{tahun}.xlsx'))
    kelola_arsip_bulanan(output_dir_excel, tahun)
    update_daftar_arsip_json(output_dir_excel)
    
    log_print(f'SELESAI GENERATE TAHUN {tahun} | Total data: {len(final_df)}')
    return len(final_df)


def kirim_telegram_aman(pesan):
    if len(pesan) > 4000: pesan = pesan[:4000] + "\n...[TERPOTONG]"
    try:
        import requests
        res = requests.post(f"https://api.telegram.org/bot{config_rahasia.BOT_TOKEN}/sendMessage", data={"chat_id": config_rahasia.CHAT_ID, "text": pesan}, timeout=10)
        if res.status_code != 200: log_print(f"🚨 GAGAL KIRIM TELEGRAM (STATUS {res.status_code}): {res.text}")
    except Exception as e: log_print(f"🚨 GAGAL KONEKSI KE TELEGRAM: {str(e)}")


if __name__ == '__main__':
    log_print("\n" + "="*55)
    log_print(f"START GENERATE PENGADAAN {get_waktu_indonesia()}")
    log_print("="*55)

    # 1. BACA ERROR DI AWAL
    daftar_error_api = []
    path_error = os.path.join(BASE_DIR, 'scripts', 'pengadaan', 'error_api_pengadaan.json')
    if os.path.exists(path_error):
        try:
            with open(path_error, 'r', encoding='utf-8') as f: daftar_error_api = json.load(f)
            os.remove(path_error)
        except: pass

    # 2. CEK TOTAL URL (Disesuaikan dengan tahun yang tidak di-skip)
    total_target = 0
    path_url = os.path.join(BASE_DIR, 'scripts', 'pengadaan', 'url_pengadaan.txt')
    if os.path.exists(path_url):
        with open(path_url, 'r', encoding='utf-8') as f:
            jumlah_url = len([line for line in f if line.strip()])
            
        tahun_diproses = 0
        for t in daftar_tahun:
            if t == tahun_n2 and os.path.exists(os.path.join(BASE_DIR, 'data', str(t), f'rekap_pengadaan_{t}.json')):
                continue
            tahun_diproses += 1
            
        total_target = jumlah_url * tahun_diproses

    # 3. LOGIKA BERHENTI JIKA GAGAL TOTAL
    if total_target > 0 and len(daftar_error_api) >= total_target:
        pesan_gagal = f"🚨 LAPORAN UPDATE (PENGADAAN) 🚨\n\n⚠️ GAGAL TOTAL DOWNLOAD API!\nTidak ada data baru. Skrip Generate dihentikan.\n\nWaktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_gagal)
        log_print("GAGAL TOTAL. Skrip berhenti.")
        sys.exit(0)

    for t in daftar_tahun: process_tahun(t)

    with open(os.path.join(BASE_DIR, "data", "last-update-pengadaan.txt"), "w", encoding='utf-8') as f:
        f.write(get_waktu_indonesia())

    # --- EKSEKUSI TEPRA ---
    log_print("\n" + "="*50)
    log_print("START GENERATE TEPRA")
    status_tepra = "✅ Data TEPRA berhasil dibuat."
    try:
        path_tepra = os.path.join(BASE_DIR, 'scripts', 'pengadaan', 'generate_tepra.py')
        res_tepra = subprocess.run([sys.executable, path_tepra], capture_output=True, text=True)
        if res_tepra.returncode == 0:
            log_print("PROSES TEPRA SUKSES\n" + res_tepra.stdout.strip())
        else:
            log_print(f"GAGAL PROSES TEPRA:\n{res_tepra.stderr}")
            status_tepra = "⚠️ Gagal membuat data TEPRA (Cek Log)."
    except Exception as e:
        log_print(f"ERROR TEPRA: {str(e)}")
        status_tepra = "⚠️ Error sistem saat eksekusi TEPRA."
    # ----------------------

    # 1. PUSH KE GITHUB
    git_sukses, pesan_git = sync_to_github()

    # --- MENGHITUNG DURASI TOTAL & AUTO DELETE ---
    durasi_str = "Tidak diketahui"
    file_start = os.path.join(BASE_DIR, 'tools', 'start_time_pengadaan.txt')
    if os.path.exists(file_start):
        try:
            with open(file_start, 'r') as f:
                waktu_mulai = float(f.read().strip())
            durasi_detik = int(time.time() - waktu_mulai)
            durasi_str = str(timedelta(seconds=durasi_detik))
            os.remove(file_start) # Auto-delete file sementara
        except: pass
    # ---------------------------------------------

    # 2. KIRIM TELEGRAM
    if len(daftar_error_api) > 0 or not git_sukses:
        pesan_ringkasan = "🚨 LAPORAN UPDATE SISTEM (PENGADAAN) 🚨\n\n"
        if len(daftar_error_api) > 0:
            teks_error = "⚠️ GAGAL DOWNLOAD API:\n" + "".join([f"{err.replace('_', ' ')}\n" for err in daftar_error_api])
            if len(teks_error) > 3500: teks_error = teks_error[:3500] + "\n... [DAFTAR ERROR DIPOTONG] ...\n"
            pesan_ringkasan += teks_error + "\n"
        pesan_ringkasan += f"📊 STATUS TEPRA: {status_tepra}\n\n🌐 STATUS GITHUB:\n{pesan_git}\n\n⏱ Durasi Total: {durasi_str}\n📅 Waktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_ringkasan)
    else:
        pesan_sukses = f"✅ UPDATE PENGADAAN BERHASIL ✅\n\nSeluruh data berhasil diolah.\n📊 STATUS TEPRA: {status_tepra}\n\n🌐 STATUS GITHUB:\n{pesan_git}\n\n⏱ Durasi Total: {durasi_str}\n📅 Waktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_sukses)
        
    log_print(f"\nPROSES SELESAI SELURUHNYA PADA {get_waktu_indonesia()}")