# ======================================================
# GENERATE DETAIL RUP (PENYEDIA & SWAKELOLA)
# ======================================================
import pandas as pd
import json
import os
import sys
import subprocess
import shutil
from datetime import datetime, timedelta, timezone
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Memastikan bisa import config_rahasia dari root folder
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import config_rahasia

warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

tahun_n  = datetime.now().year       
daftar_tahun = [tahun_n, tahun_n - 1, tahun_n - 2]

LOG_FILE = os.path.join(BASE_DIR, 'tools', 'log_detail_rup.txt')
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

def log_print(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    print(msg, **kwargs) 
    if 'end' in kwargs and kwargs['end'] == " ": return
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

def get_waktu_indonesia():
    tz_wib = timezone(timedelta(hours=7))
    sekarang = datetime.now(tz_wib)
    bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
    return f"{sekarang.day} {bulan_indo[sekarang.month]} {sekarang.year} | {sekarang.strftime('%H.%M')} WIB"

def format_tanggal_indo(tgl_str):
    if pd.isna(tgl_str) or not str(tgl_str).strip() or str(tgl_str).strip() == '-': 
        return '-'
    try:
        dt = pd.to_datetime(str(tgl_str))
        bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
        return f"{dt.day} {bulan_indo[dt.month]} {dt.year}"
    except:
        return str(tgl_str)

def format_tanggal_slash(tgl_str):
    if pd.isna(tgl_str) or not str(tgl_str).strip() or str(tgl_str).strip() == '-': 
        return '-'
    try:
        dt = pd.to_datetime(str(tgl_str))
        return dt.strftime('%d/%m/%Y')
    except:
        return str(tgl_str)        

def cek_aspek(val):
    val_str = str(val).lower()
    return "Ya" if val_str == 'true' else "Tidak"

def sync_to_github():
    log_print("\n" + "="*50)
    log_print("MENGIRIM EXCEL DETAIL RUP KE GITHUB...")
    log_print("==================================================")
    waktu_sekarang = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        subprocess.run(["git", "config", "user.name", "rizkipem-21"], cwd=BASE_DIR)
        subprocess.run(["git", "config", "user.email", "rizki.pem@gmail.com"], cwd=BASE_DIR)
        subprocess.run(["git", "add", "."], capture_output=True, text=True, cwd=BASE_DIR)
        
        commit_msg = f"Auto update Detail RUP {waktu_sekarang}"
        subprocess.run(["git", "commit", "-m", commit_msg], capture_output=True, text=True, cwd=BASE_DIR)
        res_push = subprocess.run(["git", "push"], capture_output=True, text=True, cwd=BASE_DIR)
        
        if res_push.returncode == 0:
            log_print("✅ PUSH GITHUB SUKSES!")
            return True, "✅ Push ke GitHub BERHASIL"
        else:
            log_print("⚠️ PUSH GAGAL / TIDAK ADA PERUBAHAN.")
            return False, "⚠️ Push gagal atau tidak ada perubahan."
    except Exception as e:
        log_print(f"ERROR GIT: {str(e)}")
        return False, f"❌ Error Git:\n`{str(e)}`"

def kirim_telegram_aman(pesan):
    if len(pesan) > 4000: pesan = pesan[:4000] + "\n...[TERPOTONG]"
    url = f"https://api.telegram.org/bot{config_rahasia.BOT_TOKEN}/sendMessage"
    try:
        import requests
        requests.post(url, data={"chat_id": config_rahasia.CHAT_ID, "text": pesan}, timeout=10)
    except: pass

def load_json_local(path):
    if not os.path.exists(path): return []
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
            if isinstance(data, list): return data
            if isinstance(data, dict): 
                isi = data.get('data', [])
                return isi if isi is not None else []
            return []
    except: return []

def get_file_path(data_dir, base_name, tahun):
    v1_path = os.path.join(data_dir, f"v1_{base_name}_{tahun}.json")
    legacy_path = os.path.join(data_dir, f"Legacy_{base_name}_{tahun}.json")
    return v1_path if os.path.exists(v1_path) else legacy_path

def kelola_arsip_detail(folder_path, tahun):
    if not os.path.exists(folder_path): return
    folder_arsip = os.path.join(BASE_DIR, 'arsip_lokal', 'rup', str(tahun))
    os.makedirs(folder_arsip, exist_ok=True)
    file_excel = [f for f in os.listdir(folder_path) if f.startswith('Detail_RUP') and f.endswith('.xlsx') and not f.startswith('~$')]
    
    arsip_harian = {}
    for f in file_excel:
        import re
        match = re.search(r"\((.*?)\)", f)
        if match:
            tgl = match.group(1)
            if tgl not in arsip_harian: arsip_harian[tgl] = []
            arsip_harian[tgl].append(f)
            
    list_tgl = sorted(arsip_harian.keys())
    for tgl in list_tgl[:-1]:
        for nama_file in arsip_harian[tgl]:
            path_sumber = os.path.join(folder_path, nama_file)
            path_tujuan = os.path.join(folder_arsip, nama_file)
            
            berhasil_pindah = False
            while not berhasil_pindah:
                try: 
                    shutil.move(path_sumber, path_tujuan)
                    berhasil_pindah = True
                except PermissionError:
                    import time
                    log_print(f"⚠️ Akses Ditolak: File arsip {nama_file} sedang terbuka.")
                    tutup_sukses = False
                    try:
                        import win32com.client
                        excel_app = win32com.client.GetActiveObject("Excel.Application")
                        for wb_app in excel_app.Workbooks:
                            if wb_app.Name == nama_file:
                                wb_app.Close(SaveChanges=False)
                                tutup_sukses = True
                                break
                    except Exception: pass
                    
                    if not tutup_sukses:
                        os.system(f'mshta vbscript:Execute("CreateObject(""WScript.Shell"").Popup(""File {nama_file} sedang terbuka. Tutup di Excel agar bisa diarsipkan!"", 5, ""Peringatan Excel"", 48)(window.close)")')
                    time.sleep(3)
                except Exception: 
                    berhasil_pindah = True

def update_arsip_detail_json(folder_path):
    if not os.path.exists(folder_path): return
    file_excel = [f for f in os.listdir(folder_path) if f.startswith('Detail_RUP') and f.endswith('.xlsx') and not f.startswith('~$')]
    file_excel.sort(reverse=True) 
    arsip_list = [{"nama_file": f} for f in file_excel]
    try:
        with open(os.path.join(folder_path, 'arsip_detail_rup.json'), 'w', encoding='utf-8') as f:
            json.dump(arsip_list, f, indent=4)
    except: pass

def get_dict_anggaran(filepath):
    data = load_json_local(filepath)
    if not data: return {}
    dict_anggaran = {}
    for item in data:
        kd = item.get('kd_rup')
        if not kd: continue
        if kd not in dict_anggaran: dict_anggaran[kd] = {'sd': [], 'mak': []}
        if item.get('sumber_dana'): dict_anggaran[kd]['sd'].append(str(item.get('sumber_dana')))
        if item.get('mak'): dict_anggaran[kd]['mak'].append(str(item.get('mak')))
    return dict_anggaran

def get_dict_lokasi(filepath):
    data = load_json_local(filepath)
    if not data: return {}
    dict_lok = {}
    for item in data:
        kd = item.get('kd_rup')
        if not kd: continue
        d_lok = item.get('detail_lokasi', [])
        teks = [str(l.get('detil_lokasi')) for l in d_lok if isinstance(l, dict) and l.get('detil_lokasi')]
        if teks: dict_lok[kd] = "; ".join(teks)
    return dict_lok

def process_tahun(tahun):
    log_print(f"\n--- MEMPROSES DETAIL TAHUN {tahun} ---")
    output_dir = os.path.join(BASE_DIR, "output", "rup", str(tahun))
    
    # ---------------------------------------------------------
    # LOGIKA SKIP: Melewati tahun n-1 & n-2 jika file Excel sudah ada
    # ---------------------------------------------------------
    if tahun != tahun_n and os.path.exists(output_dir):
        file_sudah_ada = any(f.startswith(f'Detail_RUP Tahun {tahun}') and f.endswith('.xlsx') for f in os.listdir(output_dir))
        if file_sudah_ada:
            log_print(f"  -> [SKIP] Excel Tahun {tahun} sudah ada (Final).")
            return 0
    # ---------------------------------------------------------
    
    data_dir = os.path.join(BASE_DIR, 'data', str(tahun))
    
    # 1. Mendapatkan PATH 6 File JSON
    p_terum = get_file_path(data_dir, "rup_paket-penyedia-terumumkan", tahun)
    p_ang   = get_file_path(data_dir, "rup_paket-anggaran-penyedia", tahun)
    p_det   = get_file_path(data_dir, "rup_paket-penyedia", tahun)
    
    s_terum = get_file_path(data_dir, "rup_paket-swakelola-terumumkan", tahun)
    s_ang   = get_file_path(data_dir, "rup_paket-anggaran-swakelola", tahun)
    s_det   = get_file_path(data_dir, "rup_paket-swakelola", tahun)

    # 2. Extract Data Tambahan (Lokasi & Anggaran)
    dict_p_lok = get_dict_lokasi(p_det)
    dict_s_lok = get_dict_lokasi(s_det)
    dict_p_ang = get_dict_anggaran(p_ang)
    dict_s_ang = get_dict_anggaran(s_ang)

    # 3. Proses List PENYEDIA
    data_p_raw = load_json_local(p_terum)
    list_p_final = []
    for item in data_p_raw:
        kd = item.get('kd_rup')
        list_p_final.append({
            'Kode RUP': kd,
            'Nama Paket': item.get('nama_paket', '-'),
            'Nama KLPD': item.get('nama_klpd', '-'),
            'Satuan Kerja': item.get('nama_satker', '-'),
            'Tahun Anggaran': item.get('tahun_anggaran', '-'),
            'Lokasi Pekerjaan': dict_p_lok.get(kd, '-'),
            'Volume Pekerjaan': item.get('volume_pekerjaan', '-'),
            'Uraian Pekerjaan': item.get('uraian_pekerjaan', item.get('urarian_pekerjaan', '-')),
            'Spesifikasi Pekerjaan': item.get('spesifikasi_pekerjaan', '-'),
            'Produk Dalam Negeri': item.get('status_pdn', '-'),
            'Usaha Kecil/Koperasi': item.get('status_ukm', '-'),
            'Aspek Ekonomi': cek_aspek(item.get('spp_aspek_ekonomi')),
            'Aspek Sosial': cek_aspek(item.get('spp_aspek_sosial')),
            'Aspek Lingkungan': cek_aspek(item.get('spp_aspek_lingkungan')),
            'Pra DIPA / DPA': item.get('status_pradipa', '-'),
            'Sumber Dana': ", ".join(dict_p_ang.get(kd, {}).get('sd', ['-'])) if kd in dict_p_ang else '-',
            'MAK': ", ".join(dict_p_ang.get(kd, {}).get('mak', ['-'])) if kd in dict_p_ang else '-',
            'Pagu': item.get('pagu', 0),
            'Jenis Pengadaan': item.get('jenis_pengadaan', '-'),
            'Metode Pemilihan': item.get('metode_pengadaan', '-'),
            'status konsolidasi': item.get('status_konsolidasi', '-'),
            'Tgl Awal Pemilihan': format_tanggal_indo(item.get('tgl_awal_pemilihan')),
            'Tgl Akhir Pemilihan': format_tanggal_indo(item.get('tgl_akhir_pemilihan')),
            'Tgl Awal Kontrak': format_tanggal_indo(item.get('tgl_awal_kontrak')),
            'Tgl Akhir Kontrak': format_tanggal_indo(item.get('tgl_akhir_kontrak')),
            'Tgl Awal Pemanfaatan': format_tanggal_indo(item.get('tgl_awal_pemanfaatan')),
            'Tgl Akhir Pemanfaatan': format_tanggal_indo(item.get('tgl_akhir_pemanfaatan')),
            'Tgl Buat Paket': format_tanggal_indo(item.get('tgl_buat_paket')),
            'Tgl Pengumuman Paket': format_tanggal_indo(item.get('tgl_pengumuman_paket')),
            'jenis paket': 'Penyedia',
            'tanggal buat': format_tanggal_slash(item.get('tgl_buat_paket')),
            'tanggal pengumuman': format_tanggal_slash(item.get('tgl_pengumuman_paket'))
        })

    # 4. Proses List SWAKELOLA
    data_s_raw = load_json_local(s_terum)
    list_s_final = []
    for item in data_s_raw:
        kd = item.get('kd_rup')
        tipe = str(item.get('tipe_swakelola', '-'))
        list_s_final.append({
            'Kode RUP': kd,
            'Nama Paket': item.get('nama_paket', '-'),
            'Nama KLPD': item.get('nama_klpd', '-'),
            'Satuan Kerja': item.get('nama_satker', '-'),
            'Tahun Anggaran': item.get('tahun_anggaran', '-'),
            'Lokasi Pekerjaan': dict_s_lok.get(kd, '-'),
            'Volume Pekerjaan': item.get('volume_pekerjaan', '-'),
            'Uraian Pekerjaan': item.get('uraian_pekerjaan', item.get('urarian_pekerjaan', '-')),
            'Spesifikasi Pekerjaan': '-',
            'Produk Dalam Negeri': '-',
            'Usaha Kecil/Koperasi': '-',
            'Aspek Ekonomi': '-',
            'Aspek Sosial': '-',
            'Aspek Lingkungan': '-',
            'Pra DIPA / DPA': '-',
            'Sumber Dana': ", ".join(dict_s_ang.get(kd, {}).get('sd', ['-'])) if kd in dict_s_ang else '-',
            'MAK': ", ".join(dict_s_ang.get(kd, {}).get('mak', ['-'])) if kd in dict_s_ang else '-',
            'Pagu': item.get('pagu', 0),
            'Jenis Pengadaan': 'Swakelola',
            'Metode Pemilihan': f"Tipe {tipe}" if tipe.isdigit() else tipe,
            'status konsolidasi': '-',
            'Tgl Awal Pemilihan': '-',
            'Tgl Akhir Pemilihan': '-',
            'Tgl Awal Kontrak': '-',
            'Tgl Akhir Kontrak': '-',
            'Tgl Awal Pemanfaatan': format_tanggal_indo(item.get('tgl_awal_pelaksanaan_kontrak')),
            'Tgl Akhir Pemanfaatan': format_tanggal_indo(item.get('tgl_akhir_pelaksanaan_kontrak')),
            'Tgl Buat Paket': format_tanggal_indo(item.get('tgl_buat_paket')),
            'Tgl Pengumuman Paket': format_tanggal_indo(item.get('tgl_pengumuman_paket')),
            'jenis paket': 'Swakelola',
            'tanggal buat': format_tanggal_slash(item.get('tgl_buat_paket')),
            'tanggal pengumuman': format_tanggal_slash(item.get('tgl_pengumuman_paket'))
        })

    # 5. Gabungkan dan Export Excel
    df_gabungan = pd.DataFrame(list_p_final + list_s_final)
    if df_gabungan.empty:
        log_print(f"Data kosong untuk tahun {tahun}.")
        return 0
        
        # PENGAMAN EXCEL: Sapu bersih karakter tersembunyi (Control Characters) dari seluruh sel
    df_gabungan = df_gabungan.replace(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', regex=True)

    # KONVERSI KE DATETIME: agar filter Excel menampilkan hierarki Tahun > Bulan
    df_gabungan['tanggal buat'] = pd.to_datetime(df_gabungan['tanggal buat'], errors='coerce', dayfirst=True)
    df_gabungan['tanggal pengumuman'] = pd.to_datetime(df_gabungan['tanggal pengumuman'], errors='coerce', dayfirst=True)

    output_dir = os.path.join(BASE_DIR, "output", "rup", str(tahun))
    os.makedirs(output_dir, exist_ok=True)
    tgl_cetak = datetime.now().strftime('%Y-%m-%d')
    nama_excel = f"Detail_RUP Tahun {tahun} ({tgl_cetak}).xlsx"
    path_excel = os.path.join(output_dir, nama_excel)

    # 6. Styling Excel dan Simpan (Dengan Fitur Auto-Close / Pop-Up)
    berhasil_simpan = False
    while not berhasil_simpan:
        try:
            df_gabungan.to_excel(path_excel, index=False, sheet_name='Detail RUP')
            
            wb = load_workbook(path_excel)
            ws = wb.active
            h_fill = PatternFill('solid', start_color='1F4E79')
            h_font = Font(name='Arial', bold=True, color='FFFFFF')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.fill, cell.font, cell.alignment = h_fill, h_font, Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col in ws.columns:
                kolom_huruf = col[0].column_letter
                if kolom_huruf in ['B', 'D', 'F', 'H', 'I', 'Q']: 
                    ws.column_dimensions[kolom_huruf].width = 40
                elif kolom_huruf == 'R': 
                    ws.column_dimensions[kolom_huruf].width = 20
                else:
                    ws.column_dimensions[kolom_huruf].width = 18

            # Deteksi posisi kolom tanggal buat & tanggal pengumuman dari header
            kolom_date = [cell.column for cell in ws[1] if cell.value in ('tanggal buat', 'tanggal pengumuman')]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    if cell.column_letter == 'R': 
                        cell.number_format = '#,##0'
                    if cell.column in kolom_date and cell.value is not None:
                        cell.number_format = 'dd/mm/yyyy'

            wb.save(path_excel)
            berhasil_simpan = True
            
        except PermissionError:
            import time
            log_print(f"⚠️ Akses Ditolak: File {nama_excel} sedang terbuka.")
            
            tutup_sukses = False
            try:
                # 1. Mencoba menutup HANYA file yang bersangkutan
                import win32com.client
                excel = win32com.client.GetActiveObject("Excel.Application")
                for wb_app in excel.Workbooks:
                    if wb_app.Name == nama_excel:
                        log_print(f"Menutup otomatis HANYA file: {nama_excel}...")
                        wb_app.Close(SaveChanges=False)
                        tutup_sukses = True
                        break
            except ImportError:
                log_print("Library pywin32 belum terinstall.")
            except Exception:
                pass 
                
            # 2. Jika gagal ditutup otomatis, munculkan Pop-Up dan tunggu
            if not tutup_sukses:
                log_print("Memunculkan pop-up peringatan ke layar...")
                pesan_popup = f"Data tidak bisa disimpan karena file {nama_excel} sedang TERBUKA. Mohon TUTUP file tersebut di Excel agar sistem bisa melanjutkan proses!"
                os.system(f'mshta vbscript:Execute("CreateObject(""WScript.Shell"").Popup(""{pesan_popup}"", 5, ""Peringatan Excel"", 48)(window.close)")')
            
            time.sleep(3) 
            log_print("Mencoba menyimpan ulang data...")

    kelola_arsip_detail(output_dir, tahun)
    update_arsip_detail_json(output_dir) # Memanggil pembuatan arsip JSON
    
    log_print(f"✅ EXCEL: {path_excel} ({len(df_gabungan)} baris)")
    return len(df_gabungan)

if __name__ == "__main__":
    tz_wib = timezone(timedelta(hours=7))
    waktu_mulai = datetime.now(tz_wib)
    
    log_print("\n==================================================")
    log_print(f"START GENERATE DETAIL RUP {get_waktu_indonesia()}")
    log_print("==================================================")

    # 1. CEK GAGAL TOTAL DOWNLOAD API (HANYA MEMBACA, TIDAK MENGHAPUS)
    path_error = os.path.join(BASE_DIR, 'scripts', 'rup', 'error_api_rup.json')
    path_url = os.path.join(BASE_DIR, 'scripts', 'rup', 'url_rup.txt')
    if os.path.exists(path_error) and os.path.exists(path_url):
        try:
            with open(path_error, 'r', encoding='utf-8') as f: err_data = json.load(f)
            with open(path_url, 'r', encoding='utf-8') as f:
                jumlah_url = len([l for l in f if l.strip() and not l.strip().startswith(('#', '='))])
            
            # Hitung tahun yang akan diproses (mengabaikan tahun yang di-skip)
            tahun_diproses = 0
            for t in daftar_tahun:
                output_dir = os.path.join(BASE_DIR, "output", "rup", str(t))
                if t != tahun_n and os.path.exists(output_dir):
                    if any(f.startswith(f'Detail_RUP Tahun {t}') and f.endswith('.xlsx') for f in os.listdir(output_dir)):
                        continue
                tahun_diproses += 1
                
            total_target = jumlah_url * tahun_diproses
            
            # Jika error memenuhi/melebihi target, matikan skrip
            if total_target > 0 and len(err_data) >= total_target:
                log_print("GAGAL TOTAL DOWNLOAD API. Skrip Detail RUP otomatis dihentikan.")
                try:
                    with open(os.path.join(BASE_DIR, 'tools', 'temp_tg_rup.txt'), 'a', encoding='utf-8') as f:
                        f.write("🔹 Detail RUP: Batal (API Error)\n")
                except: pass
                sys.exit(0)
        except Exception: pass

    total_baris = 0
    for t in daftar_tahun:
        total_baris += process_tahun(t)

    git_sukses, pesan_git = sync_to_github()
    durasi = str(datetime.now(tz_wib) - waktu_mulai).split('.')[0]

    log_print("==================================================")
    log_print(f"PROSES SELESAI | Durasi: {durasi}")
    log_print("==================================================")

    # Tulis laporan ke file sementara (Bungkam notifikasi Telegram langsung)
    temp_tg = os.path.join(BASE_DIR, 'tools', 'temp_tg_rup.txt')
    try:
        with open(temp_tg, 'a', encoding='utf-8') as f:
            f.write(f"🔹 Detail RUP: {total_baris} Baris (Durasi: {durasi})\n")
    except: pass