# ======================================================
# 2. GENERATE RUP MULTI TAHUN (JSON + EXCEL + GIT + TELEGRAM)
# ======================================================
import pandas as pd
import json
import os
import re
import sys
import subprocess
import shutil
import time
from datetime import datetime, timedelta, timezone
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import config_rahasia

warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

tahun_n  = datetime.now().year       
tahun_n1 = tahun_n - 1               
tahun_n2 = tahun_n - 2               
daftar_tahun = [tahun_n, tahun_n1, tahun_n2]

LOG_FILE = os.path.join(BASE_DIR, 'tools', 'log_rup.txt')
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

def log_print(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    print(msg, **kwargs) 
    if 'end' in kwargs and kwargs['end'] == " ": return
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

def get_waktu_indonesia():
    tz_wib = timezone(timedelta(hours=7))
    sekarang = datetime.now(tz_wib)
    bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
    return f"{sekarang.day} {bulan_indo[sekarang.month]} {sekarang.year} | {sekarang.strftime('%H.%M')} WIB"

def sync_to_github():
    log_path = LOG_FILE
    log_print("\n==================================================")
    log_print("MENGIRIM DATA RUP KE GITHUB DARI PYTHON...")
    log_print("==================================================")
    
    def tulis_log_git(teks):
        try:
            with open(log_path, "a", encoding='utf-8') as f:
                f.write(teks + "\n")
        except: pass

    waktu_sekarang = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    tulis_log_git("")
    tulis_log_git("==================================================")
    tulis_log_git(" PROSES SELESAI SELURUHNYA!")
    tulis_log_git("==================================================")

    try:
        subprocess.run(["git", "config", "user.name", "rizkipem-21"], cwd=BASE_DIR)
        subprocess.run(["git", "config", "user.email", "rizki.pem@gmail.com"], cwd=BASE_DIR)

        res_status = subprocess.run(["git", "status"], capture_output=True, text=True, cwd=BASE_DIR)
        tulis_log_git(res_status.stdout)

        res_add = subprocess.run(["git", "add", "."], capture_output=True, text=True, cwd=BASE_DIR)
        tulis_log_git(res_add.stdout)

        commit_msg = f"Auto update RUP {waktu_sekarang}"
        res_commit = subprocess.run(["git", "commit", "-m", commit_msg], capture_output=True, text=True, cwd=BASE_DIR)
        tulis_log_git(res_commit.stdout)

        res_push = subprocess.run(["git", "push"], capture_output=True, text=True, cwd=BASE_DIR)
        tulis_log_git(res_push.stdout)
        if res_push.stderr: tulis_log_git(res_push.stderr)

        status_code = res_push.returncode
        tulis_log_git(f"PUSH STATUS: {status_code}  ")
        tulis_log_git(f"SELESAI {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  ")

        if status_code == 0:
            log_print("PUSH GITHUB SUKSES!")
            return True, "✅ Push ke GitHub BERHASIL"
        else:
            raise subprocess.CalledProcessError(status_code, "git push", output=res_push.stdout, stderr=res_push.stderr)

    except subprocess.CalledProcessError as e:
        error_msg = e.stderr if e.stderr else (e.stdout if e.stdout else "Terjadi kesalahan koneksi Git.")
        log_print(f"GAGAL PUSH GITHUB: {error_msg}")
        return False, f"❌ Push ke GitHub GAGAL:\n`{error_msg.strip()}`"
    except Exception as e:
        log_print(f"ERROR SISTEM: {str(e)}")
        return False, f"❌ Terjadi kesalahan pada eksekusi Git:\n`{str(e)}`"

def kirim_telegram_aman(pesan):
    if len(pesan) > 4000:
        pesan = pesan[:4000] + "\n...[TERPOTONG]"
    url = f"https://api.telegram.org/bot{config_rahasia.BOT_TOKEN}/sendMessage"
    try:
        import requests
        res = requests.post(url, data={"chat_id": config_rahasia.CHAT_ID, "text": pesan}, timeout=10)
        if res.status_code != 200:
            log_print(f"🚨 GAGAL KIRIM TELEGRAM (STATUS {res.status_code}): {res.text}")
    except Exception as e:
        log_print(f"🚨 GAGAL KONEKSI KE TELEGRAM (JARINGAN PUTUS): {str(e)}")

def load_json_local(path):
    try:
        if not os.path.exists(path): return []
        with open(path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
            if isinstance(data, list): return data
            if isinstance(data, dict): return data.get('data', [])
            return []
    except:
        return []

def kelola_arsip_bulanan(folder_path, tahun):
    if not os.path.exists(folder_path): return
    folder_arsip_lokal = os.path.join(BASE_DIR, 'arsip_lokal', 'rup', str(tahun))
    os.makedirs(folder_arsip_lokal, exist_ok=True)
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    arsip_bulanan = {}
    for f in file_excel:
        match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", f)
        if match:
            bulan = match.group(1)[:7]
            if bulan not in arsip_bulanan: arsip_bulanan[bulan] = []
            arsip_bulanan[bulan].append((match.group(1), f))
            
    for bulan, list_file in arsip_bulanan.items():
        list_file.sort(key=lambda x: x[0])
        for tgl, nama_file in list_file[:-1]:
            try: shutil.move(os.path.join(folder_path, nama_file), os.path.join(folder_arsip_lokal, nama_file))
            except: pass

def update_daftar_arsip_json(folder_path):
    if not os.path.exists(folder_path): return
    file_excel = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    file_excel.sort(reverse=True) 
    arsip_list = [{"nama_file": f} for f in file_excel]
    try:
        with open(os.path.join(folder_path, 'daftar_arsip.json'), 'w', encoding='utf-8') as f:
            json.dump(arsip_list, f, indent=4)
    except: pass

def get_file_path(data_dir, base_name, tahun):
    v1_path = os.path.join(data_dir, f"v1_{base_name}_{tahun}.json")
    legacy_path = os.path.join(data_dir, f"Legacy_{base_name}_{tahun}.json")
    if os.path.exists(v1_path): return v1_path
    if os.path.exists(legacy_path): return legacy_path
    return legacy_path 

def process_tahun(tahun):
    log_print(f"\n--- MEMPROSES DATA TAHUN {tahun} ---")
    data_dir = os.path.join(BASE_DIR, 'data', str(tahun))
    
    s_master    = get_file_path(data_dir, "rup_master-satker", tahun)
    s_penyedia  = get_file_path(data_dir, "rup_paket-penyedia-terumumkan", tahun)
    s_swakelola = get_file_path(data_dir, "rup_paket-swakelola-terumumkan", tahun)
    s_program   = get_file_path(data_dir, "rup_program-master", tahun)
    s_struktur  = get_file_path(data_dir, "rup_struktur-anggaran-pd", tahun)

    df_master    = pd.DataFrame(load_json_local(s_master))
    df_penyedia  = pd.DataFrame(load_json_local(s_penyedia))
    df_swakelola = pd.DataFrame(load_json_local(s_swakelola))
    df_program   = pd.DataFrame(load_json_local(s_program))
    df_struktur  = pd.DataFrame(load_json_local(s_struktur))

    if df_master.empty: 
        log_print(f"Data Master kosong. Melewati tahun {tahun}.")
        return 0

    if 'kd_satker' in df_master.columns:
        df_master['kd_satker'] = pd.to_numeric(df_master['kd_satker'], errors='coerce')
    
    df_master = df_master[df_master['tahun_aktif'].astype(str).str.contains(str(tahun), na=False)]
    master_satker = df_master[['kd_satker', 'nama_satker']].drop_duplicates().dropna(subset=['kd_satker'])
    master_satker['kd_satker'] = master_satker['kd_satker'].astype(int)
    master_satker.rename(columns={'nama_satker': 'Satuan Kerja'}, inplace=True)

    for d in [df_penyedia, df_swakelola, df_program, df_struktur]:
        if not d.empty and 'kd_satker' in d.columns:
            d['kd_satker'] = pd.to_numeric(d['kd_satker'], errors='coerce')

    rup_penyedia = df_penyedia.groupby('kd_satker', as_index=False).agg(**{'RUP Penyedia': ('pagu', 'sum'), 'Paket Penyedia': ('pagu', 'count')}) if not df_penyedia.empty else pd.DataFrame(columns=['kd_satker', 'RUP Penyedia', 'Paket Penyedia'])
    rup_swakelola = df_swakelola.groupby('kd_satker', as_index=False).agg(**{'RUP Swakelola': ('pagu', 'sum'), 'Paket Swakelola': ('pagu', 'count')}) if not df_swakelola.empty else pd.DataFrame(columns=['kd_satker', 'RUP Swakelola', 'Paket Swakelola'])

    if not df_program.empty:
        kolom_ada = [c for c in ['kd_satker', 'nama_program', 'kd_program'] if c in df_program.columns]
        df_program = df_program.drop_duplicates(subset=kolom_ada)
        df_program = df_program[~df_program['nama_program'].astype(str).str.contains(r'( M$|\(M\)$)', regex=True)]
        pagu_program = df_program.groupby('kd_satker', as_index=False)['pagu_program'].sum().rename(columns={'pagu_program': 'Pagu Program'})
    else:
        pagu_program = pd.DataFrame(columns=['kd_satker', 'Pagu Program'])

    struktur = df_struktur.groupby('kd_satker', as_index=False)['belanja_pengadaan'].sum().rename(columns={'belanja_pengadaan': 'Pagu Pengadaan'}) if not df_struktur.empty else pd.DataFrame(columns=['kd_satker', 'Pagu Pengadaan'])

    df = master_satker.merge(pagu_program, on='kd_satker', how='left')
    df = df.merge(rup_penyedia, on='kd_satker', how='left')
    df = df.merge(rup_swakelola, on='kd_satker', how='left')
    df = df.merge(struktur, on='kd_satker', how='left')
    df.fillna(0, inplace=True)

    for col in ['RUP Penyedia', 'Paket Penyedia', 'RUP Swakelola', 'Paket Swakelola', 'Pagu Pengadaan']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df['Total RUP Terumumkan'] = df['RUP Penyedia'] + df['RUP Swakelola']
    df['Total Paket Terumumkan'] = df['Paket Penyedia'] + df['Paket Swakelola']
    df['Selisih RUP Terumumkan'] = df['Total RUP Terumumkan'] - df['Pagu Pengadaan']
    
    df['Persentase'] = (
        df['Total RUP Terumumkan'].astype(float) / 
        df['Pagu Pengadaan'].astype(float).replace(0, float('nan'))
    ).fillna(0) * 100

    # Menyisipkan kolom paket di sebelah kolom pagu
    df_final = df[['Satuan Kerja', 'Pagu Program', 'Pagu Pengadaan', 'Paket Penyedia', 'RUP Penyedia', 
                    'Paket Swakelola', 'RUP Swakelola', 'Total Paket Terumumkan', 'Total RUP Terumumkan', 
                    'Selisih RUP Terumumkan', 'Persentase']]
    df_final = df_final.sort_values('Satuan Kerja').reset_index(drop=True)

    output_json = os.path.join(data_dir, f"rekap_rup_{tahun}.json")
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(df_final.to_dict(orient='records'), f, ensure_ascii=False, indent=2)

    tgl = datetime.now().strftime('%Y-%m-%d')
    nama_file_history = f"Rekap RUP Tahun {tahun} ({tgl}).xlsx"
    output_history_dir = os.path.join(BASE_DIR, "output", "rup", str(tahun))
    os.makedirs(output_history_dir, exist_ok=True)
    path_history = os.path.join(output_history_dir, nama_file_history)

    df_final.to_excel(path_history, index=False, sheet_name='Rekap RUP')

    wb = load_workbook(path_history)
    ws = wb['Rekap RUP']
    h_fill = PatternFill('solid', start_color='1F4E79')
    h_font = Font(name='Arial', bold=True, color='FFFFFF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = h_fill, h_font, Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 45
    for col in ['B','C','E','G','I','J']: ws.column_dimensions[col].width = 22 # Kolom Nilai Uang
    for col in ['D','F','H']: ws.column_dimensions[col].width = 18             # Kolom Jumlah Paket
    ws.column_dimensions['K'].width = 12                                       # Kolom Persentase

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column in range(2, 11): # Kolom B sampai J (Uang & Paket)
                cell.number_format = '#,##0'
            if cell.column == 11: # Kolom K (Persentase)
                cell.number_format = '0.00"%"'

    wb.save(path_history)
    kelola_arsip_bulanan(output_history_dir, tahun)
    update_daftar_arsip_json(output_history_dir)

    log_print(f"DONE -> JSON: rekap_rup_{tahun}.json")
    log_print(f"DONE -> EXCEL: {path_history}")
    return len(df_final)

if __name__ == "__main__":
    log_print("\n==================================================")
    log_print(f"START GENERATE RUP {get_waktu_indonesia()}")
    log_print("==================================================")

    # 1. BACA ERROR DI AWAL
    daftar_error_api = []
    path_error = os.path.join(BASE_DIR, 'scripts', 'rup', 'error_api_rup.json')
    if os.path.exists(path_error):
        try:
            with open(path_error, 'r', encoding='utf-8') as f: daftar_error_api = json.load(f)
            os.remove(path_error)
        except: pass

    # 2. CEK TOTAL URL (Disesuaikan dengan tahun yang tidak di-skip)
    total_target = 0
    path_url = os.path.join(BASE_DIR, 'scripts', 'rup', 'url_rup.txt')
    if os.path.exists(path_url):
        with open(path_url, 'r', encoding='utf-8') as f:
            jumlah_url = len([line for line in f if line.strip()])
            
        tahun_diproses = 0
        for t in daftar_tahun:
            output_json_cek = os.path.join(BASE_DIR, 'data', str(t), f'rekap_rup_{t}.json')
            if t != tahun_n and os.path.exists(output_json_cek):
                continue
            tahun_diproses += 1
            
        total_target = jumlah_url * tahun_diproses

    # 3. LOGIKA BERHENTI JIKA GAGAL TOTAL
    if total_target > 0 and len(daftar_error_api) >= total_target:
        pesan_gagal = f"🚨 LAPORAN UPDATE (RUP) 🚨\n\n⚠️ GAGAL TOTAL DOWNLOAD API!\nTidak ada data baru. Skrip Generate dihentikan.\n\nWaktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_gagal)
        log_print("GAGAL TOTAL. Skrip berhenti.")
        sys.exit(0)

    # ... (lanjutkan ke kode for t in daftar_tahun: dst) ...
    total_all = 0
    for t in daftar_tahun:
        output_json_cek = os.path.join(BASE_DIR, 'data', str(t), f'rekap_rup_{t}.json')
        if t != tahun_n and os.path.exists(output_json_cek):
            log_print(f"\n[SKIP] Tahun {t} sudah ada (final).")
            continue
        total_all += process_tahun(t)
    
    with open(os.path.join(BASE_DIR, "data", "last-update-rup.txt"), "w", encoding='utf-8') as f:
        f.write(get_waktu_indonesia())

    git_sukses, pesan_git = sync_to_github()

    log_print("\n" + "="*50)
    log_print(f"PROSES SELESAI SELURUHNYA PADA {get_waktu_indonesia()}")
    log_print("==================================================")

    # --- MENGHITUNG DURASI TOTAL & AUTO DELETE ---
    durasi_str = "Tidak diketahui"
    file_start = os.path.join(BASE_DIR, 'tools', 'start_time_rup.txt')
    if os.path.exists(file_start):
        try:
            with open(file_start, 'r') as f:
                waktu_mulai = float(f.read().strip())
            durasi_detik = int(time.time() - waktu_mulai)
            durasi_str = str(timedelta(seconds=durasi_detik))
            os.remove(file_start) # Auto-delete file sementara
        except: pass
    # ---------------------------------------------

    if len(daftar_error_api) > 0 or not git_sukses:
        pesan_ringkasan = "🚨 LAPORAN UPDATE SISTEM (RUP) 🚨\n\n"
        if len(daftar_error_api) > 0:
            teks_error = "⚠️ GAGAL DOWNLOAD API:\n"
            for err in daftar_error_api:
                teks_error += f"{err.replace('_', ' ')}\n" 
            if len(teks_error) > 3500:
                teks_error = teks_error[:3500] + "\n... [DAFTAR ERROR DIPOTONG] ...\n"
            pesan_ringkasan += teks_error + "\n"
            
        pesan_ringkasan += f"🌐 STATUS GITHUB:\n{pesan_git}\n\n⏱ Durasi Total: {durasi_str}\n📅 Waktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_ringkasan)
    else:
        pesan_sukses = f"✅ UPDATE RUP BERHASIL ✅\n\nSeluruh data berhasil diolah dan sinkronisasi selesai.\n\n🌐 STATUS GITHUB:\n{pesan_git}\n\n⏱ Durasi Total: {durasi_str}\n📅 Waktu: {get_waktu_indonesia()}"
        kirim_telegram_aman(pesan_sukses)