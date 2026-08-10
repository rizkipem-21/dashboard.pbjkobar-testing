# ======================================================
# GENERATE DETAIL RUP (PENYEDIA & SWAKELOLA) - BATAS 31 MARET
# ======================================================
import pandas as pd
import json
import os
import sys
import shutil
from datetime import datetime, timedelta, timezone
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

tahun_n  = datetime.now().year       
daftar_tahun = [tahun_n, tahun_n - 1, tahun_n - 2]

LOG_FILE = os.path.join(BASE_DIR, 'tools', 'log_detail_rup_31.txt')
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

def log_print(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    print(msg, **kwargs) 
    if 'end' in kwargs and kwargs['end'] == " ": return
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

def get_waktu_indonesia():
    tz_wib = timezone(timedelta(hours=7))
    sekarang = datetime.now(tz_wib)
    bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
    return f"{sekarang.day} {bulan_indo[sekarang.month]} {sekarang.year} | {sekarang.strftime('%H.%M')} WIB"

def format_tanggal_indo(tgl_str):
    if pd.isna(tgl_str) or not str(tgl_str).strip() or str(tgl_str).strip() == '-': 
        return '-'
    try:
        dt = pd.to_datetime(str(tgl_str))
        bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}
        return f"{dt.day} {bulan_indo[dt.month]} {dt.year}"
    except:
        return str(tgl_str)

def format_tanggal_slash(tgl_str):
    if pd.isna(tgl_str) or not str(tgl_str).strip() or str(tgl_str).strip() == '-': 
        return '-'
    try:
        dt = pd.to_datetime(str(tgl_str))
        return dt.strftime('%d/%m/%Y')
    except:
        return str(tgl_str)        

def cek_aspek(val):
    val_str = str(val).lower()
    return "Ya" if val_str == 'true' else "Tidak"

def load_json_local(path):
    if not os.path.exists(path): return []
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
            if isinstance(data, list): return data
            if isinstance(data, dict): 
                isi = data.get('data')
                return isi if isi is not None else []
            return []
    except: return []

def get_file_path(data_dir, base_name, tahun):
    v1_path = os.path.join(data_dir, f"v1_{base_name}_{tahun}.json")
    legacy_path = os.path.join(data_dir, f"Legacy_{base_name}_{tahun}.json")
    return v1_path if os.path.exists(v1_path) else legacy_path

def kelola_arsip_detail(folder_path, tahun):
    if not os.path.exists(folder_path): return
    folder_arsip = os.path.join(BASE_DIR, 'arsip_lokal', 'rup', str(tahun))
    os.makedirs(folder_arsip, exist_ok=True)
    file_excel = [f for f in os.listdir(folder_path) if f.startswith('Detail_RUP_31') and f.endswith('.xlsx')]
    
    arsip_harian = {}
    for f in file_excel:
        import re
        match = re.search(r"\((.*?)\)", f)
        if match:
            tgl = match.group(1)
            if tgl not in arsip_harian: arsip_harian[tgl] = []
            arsip_harian[tgl].append(f)
            
    list_tgl = sorted(arsip_harian.keys())
    for tgl in list_tgl[:-1]:
        for nama_file in arsip_harian[tgl]:
            try: shutil.move(os.path.join(folder_path, nama_file), os.path.join(folder_arsip, nama_file))
            except: pass

def update_arsip_detail_json(folder_path):
    if not os.path.exists(folder_path): return
    file_excel = [f for f in os.listdir(folder_path) if f.startswith('Detail_RUP_31') and f.endswith('.xlsx')]
    file_excel.sort(reverse=True) 
    arsip_list = [{"nama_file": f} for f in file_excel]
    try:
        with open(os.path.join(folder_path, 'arsip_detail_rup_31.json'), 'w', encoding='utf-8') as f:
            json.dump(arsip_list, f, indent=4)
    except: pass

def get_dict_anggaran(filepath):
    data = load_json_local(filepath)
    dict_anggaran = {}
    if not data: return dict_anggaran
    for item in data:
        kd = item.get('kd_rup')
        if not kd: continue
        if kd not in dict_anggaran: dict_anggaran[kd] = {'sd': [], 'mak': []}
        if item.get('sumber_dana'): dict_anggaran[kd]['sd'].append(str(item.get('sumber_dana')))
        if item.get('mak'): dict_anggaran[kd]['mak'].append(str(item.get('mak')))
    return dict_anggaran

def get_dict_lokasi(filepath):
    data = load_json_local(filepath)
    dict_lok = {}
    if not data: return dict_lok
    for item in data:
        kd = item.get('kd_rup')
        if not kd: continue
        d_lok = item.get('detail_lokasi', [])
        if not d_lok: continue
        teks = [str(l.get('detil_lokasi')) for l in d_lok if isinstance(l, dict) and l.get('detil_lokasi')]
        if teks: dict_lok[kd] = "; ".join(teks)
    return dict_lok

# ======================================================
# FUNGSI PELACAKAN SILSILAH TANGGAL PENGUMUMAN
# ======================================================
def get_dict_tanggal_global(p_det, s_det):
    d_tgl = {}
    for path in [p_det, s_det]:
        data = load_json_local(path)
        if not data: continue
        for item in data:
            kd = item.get('kd_rup')
            tgl = item.get('tgl_pengumuman_paket')
            if kd and tgl: d_tgl[str(kd)] = str(tgl)
    return d_tgl

def trace_tanggal_sah(kd_rup, awal_tgl_str, tahun, dict_ku, dict_mk, dict_tgl):
    batas_waktu = pd.to_datetime(f"{tahun}-03-31 23:59:59", utc=True)
    curr_kd = str(kd_rup)
    curr_tgl_str = awal_tgl_str
    visited = set()

    while True:
        try: dt = pd.to_datetime(curr_tgl_str, utc=True) if curr_tgl_str else None
        except: dt = None
        
        # Lolos pengecekan: Jika format sah dan <= 31 Maret
        if dt is not None and dt <= batas_waktu:
            return curr_tgl_str
            
        # Cegah Infinite Loop
        if curr_kd in visited: break
        visited.add(curr_kd)
        
        # Cek silsilah ID di Kaji Ulang atau Manual Konsolidasi
        kd_lama = dict_ku.get(curr_kd) or dict_mk.get(curr_kd)
        if not kd_lama: break # Silsilah mentok
            
        tgl_lama_str = dict_tgl.get(kd_lama)
        if not tgl_lama_str: break # Data lama tidak ada di memori paket
            
        curr_kd = kd_lama
        curr_tgl_str = tgl_lama_str
        
    return curr_tgl_str

# ======================================================
# PROSES UTAMA
# ======================================================

# Load Master Manual Konsolidasi Sekali di Awal
path_mk = os.path.join(BASE_DIR, 'data_master', 'manual_konsolidasi.json')
dict_mk_global = {}
if os.path.exists(path_mk):
    try:
        with open(path_mk, 'r', encoding='utf-8') as f:
            data_mk = json.load(f)
            for item in data_mk:
                b = item.get('kode_konsol_baru')
                l = item.get('kode_konsol_lama')
                if b and l: dict_mk_global[str(b)] = str(l)
    except: pass

def process_tahun(tahun):
    log_print(f"\n--- MEMPROSES DETAIL (BATAS 31 MARET) TAHUN {tahun} ---")
    data_dir = os.path.join(BASE_DIR, 'data', str(tahun))
    
    p_terum = get_file_path(data_dir, "rup_paket-penyedia-terumumkan", tahun)
    p_ang   = get_file_path(data_dir, "rup_paket-anggaran-penyedia", tahun)
    p_det   = get_file_path(data_dir, "rup_paket-penyedia", tahun)
    
    s_terum = get_file_path(data_dir, "rup_paket-swakelola-terumumkan", tahun)
    s_ang   = get_file_path(data_dir, "rup_paket-anggaran-swakelola", tahun)
    s_det   = get_file_path(data_dir, "rup_paket-swakelola", tahun)
    
    # Load History Kaji Ulang
    p_kaji = get_file_path(data_dir, "rup_history-kaji-ulang", tahun)
    data_kaji = load_json_local(p_kaji)
    dict_ku_lokal = {}
    if data_kaji:
        for item in data_kaji:
            baru = item.get('kd_rup_baru')
            lama = item.get('kd_rup_lama')
            if baru and lama: dict_ku_lokal[str(baru)] = str(lama)

    # Ekstrak Data Tambahan & Kamus
    dict_p_lok = get_dict_lokasi(p_det)
    dict_s_lok = get_dict_lokasi(s_det)
    dict_p_ang = get_dict_anggaran(p_ang)
    dict_s_ang = get_dict_anggaran(s_ang)
    dict_tgl_all = get_dict_tanggal_global(p_det, s_det)

    # 3. Proses List PENYEDIA
    data_p_raw = load_json_local(p_terum)
    list_p_final = []
    for item in data_p_raw:
        kd = item.get('kd_rup')
        tgl_pengumuman_valid = trace_tanggal_sah(kd, item.get('tgl_pengumuman_paket'), tahun, dict_ku_lokal, dict_mk_global, dict_tgl_all)
        
        list_p_final.append({
            'Kode RUP': kd,
            'Nama Paket': item.get('nama_paket', '-'),
            'Nama KLPD': item.get('nama_klpd', '-'),
            'Satuan Kerja': item.get('nama_satker', '-'),
            'Tahun Anggaran': item.get('tahun_anggaran', '-'),
            'Lokasi Pekerjaan': dict_p_lok.get(kd, '-'),
            'Volume Pekerjaan': item.get('volume_pekerjaan', '-'),
            'Uraian Pekerjaan': item.get('uraian_pekerjaan', item.get('urarian_pekerjaan', '-')),
            'Spesifikasi Pekerjaan': item.get('spesifikasi_pekerjaan', '-'),
            'Produk Dalam Negeri': item.get('status_pdn', '-'),
            'Usaha Kecil/Koperasi': item.get('status_ukm', '-'),
            'Aspek Ekonomi': cek_aspek(item.get('spp_aspek_ekonomi')),
            'Aspek Sosial': cek_aspek(item.get('spp_aspek_sosial')),
            'Aspek Lingkungan': cek_aspek(item.get('spp_aspek_lingkungan')),
            'Pra DIPA / DPA': item.get('status_pradipa', '-'),
            'Sumber Dana': ", ".join(dict_p_ang.get(kd, {}).get('sd', ['-'])) if kd in dict_p_ang else '-',
            'MAK': ", ".join(dict_p_ang.get(kd, {}).get('mak', ['-'])) if kd in dict_p_ang else '-',
            'Pagu': item.get('pagu', 0),
            'Jenis Pengadaan': item.get('jenis_pengadaan', '-'),
            'Metode Pemilihan': item.get('metode_pengadaan', '-'),
            'status konsolidasi': item.get('status_konsolidasi', '-'),
            'Tgl Awal Pemilihan': format_tanggal_indo(item.get('tgl_awal_pemilihan')),
            'Tgl Akhir Pemilihan': format_tanggal_indo(item.get('tgl_akhir_pemilihan')),
            'Tgl Awal Kontrak': format_tanggal_indo(item.get('tgl_awal_kontrak')),
            'Tgl Akhir Kontrak': format_tanggal_indo(item.get('tgl_akhir_kontrak')),
            'Tgl Awal Pemanfaatan': format_tanggal_indo(item.get('tgl_awal_pemanfaatan')),
            'Tgl Akhir Pemanfaatan': format_tanggal_indo(item.get('tgl_akhir_pemanfaatan')),
            'Tgl Buat Paket': format_tanggal_indo(item.get('tgl_buat_paket')),
            'Tgl Pengumuman Paket': format_tanggal_indo(tgl_pengumuman_valid),
            'jenis paket': 'Penyedia',
            'tanggal buat': format_tanggal_slash(item.get('tgl_buat_paket')),
            'tanggal pengumuman': format_tanggal_slash(tgl_pengumuman_valid)
        })

    # 4. Proses List SWAKELOLA
    data_s_raw = load_json_local(s_terum)
    list_s_final = []
    for item in data_s_raw:
        kd = item.get('kd_rup')
        tipe = str(item.get('tipe_swakelola', '-'))
        tgl_pengumuman_valid = trace_tanggal_sah(kd, item.get('tgl_pengumuman_paket'), tahun, dict_ku_lokal, dict_mk_global, dict_tgl_all)
        
        list_s_final.append({
            'Kode RUP': kd,
            'Nama Paket': item.get('nama_paket', '-'),
            'Nama KLPD': item.get('nama_klpd', '-'),
            'Satuan Kerja': item.get('nama_satker', '-'),
            'Tahun Anggaran': item.get('tahun_anggaran', '-'),
            'Lokasi Pekerjaan': dict_s_lok.get(kd, '-'),
            'Volume Pekerjaan': item.get('volume_pekerjaan', '-'),
            'Uraian Pekerjaan': item.get('uraian_pekerjaan', item.get('urarian_pekerjaan', '-')),
            'Spesifikasi Pekerjaan': '-',
            'Produk Dalam Negeri': '-',
            'Usaha Kecil/Koperasi': '-',
            'Aspek Ekonomi': '-',
            'Aspek Sosial': '-',
            'Aspek Lingkungan': '-',
            'Pra DIPA / DPA': '-',
            'Sumber Dana': ", ".join(dict_s_ang.get(kd, {}).get('sd', ['-'])) if kd in dict_s_ang else '-',
            'MAK': ", ".join(dict_s_ang.get(kd, {}).get('mak', ['-'])) if kd in dict_s_ang else '-',
            'Pagu': item.get('pagu', 0),
            'Jenis Pengadaan': 'Swakelola',
            'Metode Pemilihan': f"Tipe {tipe}" if tipe.isdigit() else tipe,
            'status konsolidasi': '-',
            'Tgl Awal Pemilihan': '-',
            'Tgl Akhir Pemilihan': '-',
            'Tgl Awal Kontrak': '-',
            'Tgl Akhir Kontrak': '-',
            'Tgl Awal Pemanfaatan': format_tanggal_indo(item.get('tgl_awal_pelaksanaan_kontrak')),
            'Tgl Akhir Pemanfaatan': format_tanggal_indo(item.get('tgl_akhir_pelaksanaan_kontrak')),
            'Tgl Buat Paket': format_tanggal_indo(item.get('tgl_buat_paket')),
            'Tgl Pengumuman Paket': format_tanggal_indo(tgl_pengumuman_valid),
            'jenis paket': 'Swakelola',
            'tanggal buat': format_tanggal_slash(item.get('tgl_buat_paket')),
            'tanggal pengumuman': format_tanggal_slash(tgl_pengumuman_valid)
        })

    # 5. Gabungkan dan Export Excel
    df_gabungan = pd.DataFrame(list_p_final + list_s_final)
    if df_gabungan.empty:
        log_print(f"Data kosong untuk tahun {tahun}.")
        return 0
        
    df_gabungan = df_gabungan.replace(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', regex=True)
    df_gabungan['tanggal buat'] = pd.to_datetime(df_gabungan['tanggal buat'], errors='coerce', dayfirst=True)
    df_gabungan['tanggal pengumuman'] = pd.to_datetime(df_gabungan['tanggal pengumuman'], errors='coerce', dayfirst=True)

    output_dir = os.path.join(BASE_DIR, "output", "rup", str(tahun))
    os.makedirs(output_dir, exist_ok=True)
    tgl_cetak = datetime.now().strftime('%Y-%m-%d')
    
    # Penamaan khusus untuk file ini agar tidak bentrok
    nama_excel = f"Detail_RUP_31 Tahun {tahun} ({tgl_cetak}).xlsx"
    path_excel = os.path.join(output_dir, nama_excel)

    df_gabungan.to_excel(path_excel, index=False, sheet_name='Detail RUP 31')
    
    # 6. Styling Excel
    wb = load_workbook(path_excel)
    ws = wb.active
    h_fill = PatternFill('solid', start_color='1F4E79')
    h_font = Font(name='Arial', bold=True, color='FFFFFF')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = h_fill, h_font, Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col in ws.columns:
        kolom_huruf = col[0].column_letter
        if kolom_huruf in ['B', 'D', 'F', 'H', 'I', 'Q']: 
            ws.column_dimensions[kolom_huruf].width = 40
        elif kolom_huruf == 'R': 
            ws.column_dimensions[kolom_huruf].width = 20
        else:
            ws.column_dimensions[kolom_huruf].width = 18

    kolom_date = [cell.column for cell in ws[1] if cell.value in ('tanggal buat', 'tanggal pengumuman')]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column_letter == 'R': 
                cell.number_format = '#,##0'
            if cell.column in kolom_date and cell.value is not None:
                cell.number_format = 'dd/mm/yyyy'

    wb.save(path_excel)
    kelola_arsip_detail(output_dir, tahun)
    update_arsip_detail_json(output_dir) 
    
    log_print(f"✅ EXCEL: {path_excel} ({len(df_gabungan)} baris)")
    return len(df_gabungan)

if __name__ == "__main__":
    tz_wib = timezone(timedelta(hours=7))
    waktu_mulai = datetime.now(tz_wib)
    
    log_print("\n==================================================")
    log_print(f"START GENERATE DETAIL RUP (BATAS 31 MARET) {get_waktu_indonesia()}")
    log_print("==================================================")

    total_baris = 0
    for t in daftar_tahun:
        output_dir = os.path.join(BASE_DIR, "output", "rup", str(t))
        excel_sudah_ada = False
        if os.path.exists(output_dir):
            for f in os.listdir(output_dir):
                if f.startswith('Detail_RUP_31') and f.endswith('.xlsx'):
                    excel_sudah_ada = True
                    break
                    
        if t != tahun_n and excel_sudah_ada:
            log_print(f"\n[SKIP] Tahun {t} sudah ada (final).")
            continue
            
        total_baris += process_tahun(t)

    durasi = str(datetime.now(tz_wib) - waktu_mulai).split('.')[0]

    log_print("==================================================")
    log_print(f"PROSES SELESAI SELURUHNYA | Durasi: {durasi}")
    log_print(f"Total Baris Diekstrak: {total_baris}")
    log_print("==================================================")